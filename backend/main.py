from fastapi import FastAPI, Depends, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from sqlalchemy.orm import Session
from sqlalchemy import text as sql_text
from typing import Optional, List
from pydantic import BaseModel
import csv
import hashlib
import io
import json
import re
import uuid
import os
from datetime import datetime

from models import Report, CaseLinkage, ReportStage, ResearchNote, init_db, get_db
from schemas import ReportCreate, ReportUpdate, ReportOut, SuggestRequest, StageCreate, StageUpdate, StageOut, StageReorderItem
from ai import get_ai_suggestions, parse_bulletin
from parser import parse_bulletin_rules
from similarity import compute_similarity, STOPWORDS


def _narrative_hash(raw: str) -> str:
    normalized = re.sub(r'\s+', ' ', raw.strip()).lower()
    return hashlib.sha256(normalized.encode('utf-8')).hexdigest()


def _narrative_similarity(text_a: str, text_b: str) -> float:
    """
    Returns the higher of Jaccard and overlap-coefficient similarity.

    Overlap coefficient = |A ∩ B| / min(|A|, |B|).

    This handles the common cross-format case where one text (e.g. an Excel
    synopsis cell) is topically a subset of the other (e.g. a full PDF bulletin
    entry that includes headers, dates, labels, and the same synopsis text).
    In that case Jaccard is dragged down by the extra bulletin words, but the
    overlap coefficient stays high because the smaller set is mostly covered.
    """
    wa = {w.lower() for w in text_a.split() if len(w) > 2 and w.lower() not in STOPWORDS}
    wb = {w.lower() for w in text_b.split() if len(w) > 2 and w.lower() not in STOPWORDS}
    if not wa or not wb:
        return 0.0
    inter = len(wa & wb)
    jaccard = inter / len(wa | wb)
    overlap = inter / min(len(wa), len(wb))
    return max(jaccard, overlap)


app = FastAPI(title="Red Light Alert API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:4173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def startup():
    init_db()


# ── Reports CRUD ──────────────────────────────────────────────────────────────

_VALID_NLP_PATTERNS = {
    'condom_refusal', 'payment_dispute', 'bait_and_switch',
    'rapid_escalation', 'weapon_present', 'multi_suspect', 'online_lure',
    'drugging_intoxication', 'confinement',
}


@app.get("/reports", response_model=list[ReportOut])
def list_reports(
    coding_status: Optional[str] = None,
    city: Optional[str] = None,
    coercion_present: Optional[str] = None,
    movement_present: Optional[str] = None,
    physical_force: Optional[str] = None,
    sexual_assault: Optional[str] = None,
    threats_present: Optional[str] = None,
    vehicle_present: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    search: Optional[str] = None,
    # NLP signal filters (drill-down from Analysis page)
    nlp_coercion: Optional[str] = None,      # "1" = strong only, "2" = strong+possible
    nlp_physical: Optional[str] = None,
    nlp_sexual: Optional[str] = None,
    nlp_movement: Optional[str] = None,
    nlp_weapon: Optional[str] = None,
    nlp_escalation_min: Optional[str] = None, # minimum escalation score (1–5)
    nlp_pattern: Optional[str] = None,        # named pattern, e.g. "weapon_present"
    cross_city_movement: Optional[str] = None, # yes/no/unclear
    db: Session = Depends(get_db),
):
    q = db.query(Report)
    if coding_status:
        q = q.filter(Report.coding_status == coding_status)
    if city:
        # Search across all city fields — legacy summary + stage-specific
        q = q.filter(
            Report.city.ilike(f"%{city}%") |
            Report.initial_contact_city.ilike(f"%{city}%") |
            Report.incident_city.ilike(f"%{city}%") |
            Report.destination_city.ilike(f"%{city}%")
        )
    if coercion_present:
        q = q.filter(Report.coercion_present == coercion_present)
    if movement_present:
        q = q.filter(Report.movement_present == movement_present)
    if physical_force:
        q = q.filter(Report.physical_force == physical_force)
    if sexual_assault:
        q = q.filter(Report.sexual_assault == sexual_assault)
    if threats_present:
        q = q.filter(Report.threats_present == threats_present)
    if vehicle_present:
        q = q.filter(Report.vehicle_present == vehicle_present)
    if date_from:
        q = q.filter(Report.incident_date >= date_from)
    if date_to:
        q = q.filter(Report.incident_date <= date_to)
    if search:
        q = q.filter(
            Report.raw_narrative.ilike(f"%{search}%") |
            Report.suspect_description_text.ilike(f"%{search}%") |
            Report.vehicle_make.ilike(f"%{search}%") |
            Report.plate_partial.ilike(f"%{search}%")
        )
    # NLP signal filters — JSON path extraction (SQLite json_extract)
    _nlp_rank_fields = [
        (nlp_coercion,  '$.nlp.coercion_rank'),
        (nlp_physical,  '$.nlp.physical_rank'),
        (nlp_sexual,    '$.nlp.sexual_rank'),
        (nlp_movement,  '$.nlp.movement_rank'),
        (nlp_weapon,    '$.nlp.weapon_rank'),
    ]
    for rank_param, json_path in _nlp_rank_fields:
        if rank_param:
            try:
                rank_val = int(rank_param)
            except ValueError:
                continue
            q = q.filter(sql_text(
                f"json_extract(ai_suggestions, '{json_path}') IS NOT NULL AND "
                f"CAST(json_extract(ai_suggestions, '{json_path}') AS INTEGER) <= {rank_val}"
            ))
    if nlp_escalation_min:
        try:
            min_score = int(nlp_escalation_min)
            q = q.filter(sql_text(
                f"json_extract(ai_suggestions, '$.nlp.escalation.score') IS NOT NULL AND "
                f"CAST(json_extract(ai_suggestions, '$.nlp.escalation.score') AS INTEGER) >= {min_score}"
            ))
        except ValueError:
            pass
    if nlp_pattern and nlp_pattern in _VALID_NLP_PATTERNS:
        q = q.filter(sql_text(
            f"ai_suggestions LIKE '%\"{nlp_pattern}\"%'"
        ))
    if cross_city_movement:
        q = q.filter(Report.cross_city_movement == cross_city_movement)
    return q.order_by(Report.created_at.desc()).all()


@app.post("/reports", response_model=ReportOut)
def create_report(data: ReportCreate, db: Session = Depends(get_db)):
    h = _narrative_hash(data.raw_narrative)
    existing = db.query(Report).filter(Report.narrative_hash == h).first()
    if existing:
        raise HTTPException(status_code=409,
            detail=f"Duplicate narrative — already stored as {existing.report_id}")
    report_id = f"RLA-{datetime.utcnow().strftime('%Y%m%d')}-{str(uuid.uuid4())[:6].upper()}"
    report = Report(report_id=report_id, narrative_hash=h, **data.model_dump())
    report.audit_log = [{"ts": datetime.utcnow().isoformat(), "action": "created", "by": data.analyst_name or "system"}]
    db.add(report)
    db.commit()
    db.refresh(report)
    return report


@app.get("/reports/{report_id}", response_model=ReportOut)
def get_report(report_id: str, db: Session = Depends(get_db)):
    r = db.query(Report).filter(Report.report_id == report_id).first()
    if not r:
        raise HTTPException(404, "Report not found")
    return r


@app.patch("/reports/{report_id}", response_model=ReportOut)
def update_report(report_id: str, data: ReportUpdate, db: Session = Depends(get_db)):
    r = db.query(Report).filter(Report.report_id == report_id).first()
    if not r:
        raise HTTPException(404, "Report not found")

    update_data = data.model_dump(exclude_unset=True)
    log = r.audit_log or []

    for key, val in update_data.items():
        old = getattr(r, key, None)
        if old != val:
            log.append({
                "ts": datetime.utcnow().isoformat(),
                "action": "updated",
                "field": key,
                "from": str(old),
                "to": str(val),
                "by": data.analyst_name or "system",
            })
            setattr(r, key, val)

    r.audit_log = log
    r.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(r)
    return r


@app.delete("/reports/{report_id}")
def delete_report(report_id: str, db: Session = Depends(get_db)):
    r = db.query(Report).filter(Report.report_id == report_id).first()
    if not r:
        raise HTTPException(404, "Report not found")
    db.delete(r)
    db.commit()
    return {"ok": True}


class BulkDeleteRequest(BaseModel):
    report_ids: List[str]


@app.post("/reports/bulk-delete")
def bulk_delete_reports(data: BulkDeleteRequest, db: Session = Depends(get_db)):
    if not data.report_ids:
        return {"ok": True, "deleted": 0}
    deleted = db.query(Report).filter(Report.report_id.in_(data.report_ids)).delete(synchronize_session=False)
    db.commit()
    return {"ok": True, "deleted": deleted}


# ── AI suggestions ────────────────────────────────────────────────────────────

@app.post("/suggest")
async def suggest(data: SuggestRequest):
    suggestions = await get_ai_suggestions(data.narrative)
    return suggestions


@app.post("/reports/{report_id}/analyze")
def analyze_report(report_id: str, db: Session = Depends(get_db)):
    """
    Run spaCy NLP analysis + weather fetch on a report and update ai_suggestions.
    Weather is only fetched when the report has an exact incident_date and city.
    """
    from nlp_analysis import analyze_narrative
    from weather import fetch_weather
    r = db.query(Report).filter(Report.report_id == report_id).first()
    if not r:
        raise HTTPException(404, "Report not found")

    result = analyze_narrative(r.raw_narrative or "")

    # Resolve hour from incident_time_exact (HH:MM) for hourly weather lookup
    weather_hour: int | None = None
    if r.incident_time_exact:
        try:
            weather_hour = int(r.incident_time_exact.split(":")[0])
        except (ValueError, IndexError):
            pass

    # Fetch historical weather when we have a real date (not vague/range) and city
    weather_data: dict = {}
    date_certainty = (result.get("nlp", {}) or {}).get("date_certainty", "")
    if (
        r.incident_date
        and len(r.incident_date) == 10
        and (r.incident_city or r.city)
        and date_certainty not in ("vague", "range")
    ):
        weather_city = r.incident_city or r.city
        weather_data = fetch_weather(r.incident_date, weather_city, hour=weather_hour)
        if "error" in weather_data:
            weather_data = {"error": weather_data["error"]}

    if weather_data:
        result["weather"] = weather_data

    # Stamp provenance so the frontend can verify the NLP data belongs to this report
    if "nlp" in result and isinstance(result["nlp"], dict):
        result["nlp"]["_source_report_id"] = report_id
        result["nlp"]["_analyzed_at"] = datetime.utcnow().isoformat()

    # Merge into existing ai_suggestions, preserving any other keys.
    # Use a new dict so SQLAlchemy detects the JSON column as changed.
    r.ai_suggestions = {**(r.ai_suggestions or {}), **result}
    r.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(r)
    return {"ok": True, "ai_suggestions": r.ai_suggestions}


# ── Batch NLP re-run ─────────────────────────────────────────────────────────

@app.post("/reports/batch-analyze")
def batch_analyze(db: Session = Depends(get_db)):
    """
    Run spaCy NLP analysis on every report that has no nlp data yet.
    Skips reports that already have ai_suggestions["nlp"] populated with coercion_rank.
    Returns a count of reports processed and whether spaCy is available.
    """
    from nlp_analysis import analyze_narrative, _nlp as _spacy_model
    # Return early if spaCy model not loaded — no point iterating all reports
    if _spacy_model is None:
        return {"ok": True, "processed": 0, "nlp_available": False}
    reports = db.query(Report).all()
    processed = 0
    for r in reports:
        existing = r.ai_suggestions or {}
        # Skip only if a real NLP analysis was completed (coercion_rank is set).
        # Records that only have date_certainty (from a failed spaCy run) are re-analyzed.
        if existing.get("nlp", {}).get("coercion_rank") is not None:
            continue
        if not r.raw_narrative or not r.raw_narrative.strip():
            continue
        result = analyze_narrative(r.raw_narrative)
        # Use a new dict so SQLAlchemy detects the JSON column as changed.
        r.ai_suggestions = {**existing, **result}
        r.updated_at = datetime.utcnow()
        processed += 1
    db.commit()
    return {"ok": True, "processed": processed, "nlp_available": True}


# ── Bulletin import ───────────────────────────────────────────────────────────

_ATTACHMENTS_DIR = os.path.join(os.path.dirname(__file__), "..", "attachments")
os.makedirs(_ATTACHMENTS_DIR, exist_ok=True)


@app.post("/parse-bulletin")
async def parse_bulletin_endpoint(file: UploadFile = File(...)):
    """
    Parse a Red Light Alert bulletin PDF into individual incident records.
    Uses AI (Claude) if ANTHROPIC_API_KEY is set, otherwise falls back to
    rule-based PDF column detection + regex extraction.
    Returns a session_id that links to the stored source PDF.
    """
    content = await file.read()
    filename = file.filename or ""
    has_api_key = bool(os.environ.get("ANTHROPIC_API_KEY"))

    # Generate a session ID and persist the source PDF for all methods
    session_id = str(uuid.uuid4())
    if filename.lower().endswith(".pdf"):
        pdf_path = os.path.join(_ATTACHMENTS_DIR, f"{session_id}.pdf")
        with open(pdf_path, "wb") as f:
            f.write(content)

    if filename.lower().endswith(".pdf"):
        # Always extract full text for provenance
        import tempfile, pdfplumber
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        try:
            text_pages = []
            with pdfplumber.open(tmp_path) as pdf:
                for page in pdf.pages:
                    t = page.extract_text()
                    if t:
                        text_pages.append(t)
            bulletin_text = "\n\n".join(text_pages)
        finally:
            os.unlink(tmp_path)

        if has_api_key:
            try:
                incidents = await parse_bulletin(bulletin_text)
                for inc in incidents:
                    inc["_bulletin_text"] = bulletin_text
                    inc["_session_id"] = session_id
                return {"incidents": incidents, "total": len(incidents), "method": "ai", "session_id": session_id}
            except ValueError as e:
                raise HTTPException(400, str(e))
        else:
            # Rule-based path: use PDF column detection + regex
            incidents = parse_bulletin_rules(content)
            for inc in incidents:
                inc["_bulletin_text"] = bulletin_text
                inc["_session_id"] = session_id
            return {"incidents": incidents, "total": len(incidents), "method": "rules", "session_id": session_id}
    else:
        # Plain text — always use AI if available, else return error
        bulletin_text = content.decode("utf-8", errors="replace")
        if not has_api_key:
            raise HTTPException(400, "Plain text import requires an AI API key. Upload a PDF instead, or add your ANTHROPIC_API_KEY.")
        try:
            incidents = await parse_bulletin(bulletin_text)
            for inc in incidents:
                inc["_bulletin_text"] = bulletin_text
                inc["_session_id"] = session_id
            return {"incidents": incidents, "total": len(incidents), "method": "ai", "session_id": session_id}
        except ValueError as e:
            raise HTTPException(400, str(e))


@app.get("/attachments/{session_id}")
def get_attachment(session_id: str):
    """Serve a stored source PDF by its session ID."""
    # Sanitise: only allow UUID-shaped filenames
    if not re.fullmatch(r"[0-9a-f\-]{36}", session_id):
        raise HTTPException(400, "Invalid session ID")
    pdf_path = os.path.join(_ATTACHMENTS_DIR, f"{session_id}.pdf")
    if not os.path.isfile(pdf_path):
        raise HTTPException(404, "Attachment not found")
    return FileResponse(pdf_path, media_type="application/pdf", filename="source_bulletin.pdf")


@app.post("/parse-excel")
async def parse_excel_endpoint(file: UploadFile = File(...)):
    """
    Parse an Excel dataset (same column layout as DTE DATASET for QGIS.xlsx)
    into individual incident records for preview before bulk-save.
    """
    import tempfile
    import openpyxl
    from import_excel import (
        parse_date, parse_time, parse_vehicle, parse_locations,
        clean_city, extract_neighbourhood, parse_suspect_count, parse_gender,
    )

    content = await file.read()

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(content)
        tmp_path = tmp.name

    try:
        wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
        ws = wb["All Incidents"] if "All Incidents" in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    finally:
        os.unlink(tmp_path)

    header_row = rows[0] if rows else ()
    data_rows = rows[1:]  # skip header row
    incidents = []

    for row in data_rows:
        def _cell(idx):
            return row[idx] if len(row) > idx else None

        # Build full-row source text for the Source Immutable panel
        row_source_parts = []
        for col_idx, cell_val in enumerate(row):
            header = (
                str(header_row[col_idx]).strip()
                if header_row and col_idx < len(header_row) and header_row[col_idx] is not None
                else f"Column {col_idx + 1}"
            )
            val_str = str(cell_val).strip() if cell_val is not None else ""
            row_source_parts.append(f"{header}: {val_str}")
        row_source_text = "\n".join(row_source_parts)

        synopsis_raw      = _cell(9)
        synopsis = str(synopsis_raw).strip() if synopsis_raw else ""
        if not synopsis or synopsis.lower() == "none":
            continue

        incident_date_raw = _cell(0)
        time_raw          = _cell(2)
        date_reported_raw = _cell(3)
        city_raw          = _cell(4)
        location_raw      = _cell(5)
        coords_raw        = _cell(6)
        description_raw   = _cell(7)
        vehicle_raw       = _cell(8)

        description   = str(description_raw).strip() if description_raw else ""
        vehicle_text  = str(vehicle_raw).strip()     if vehicle_raw     else ""
        location_text = str(location_raw).strip()    if location_raw    else ""

        veh  = parse_vehicle(vehicle_text)
        locs = parse_locations(location_text, synopsis)

        lat_initial = lon_initial = None
        if coords_raw and str(coords_raw).strip() not in ("None", ""):
            raw_str = str(coords_raw).strip().strip("()")
            parts = raw_str.split(",")
            if len(parts) == 2:
                try:
                    lat_initial = float(parts[0].strip())
                    lon_initial = float(parts[1].strip())
                except ValueError:
                    pass

        incidents.append({
            "raw_narrative":              synopsis,
            "entry_type":                 "incident",
            "bulletin_date":              parse_date(date_reported_raw),
            "source_organization":        "Red Light Alert",
            "incident_date":              parse_date(incident_date_raw),
            "date_reported":              parse_date(date_reported_raw),
            "city":                       clean_city(city_raw),
            "neighbourhood":              extract_neighbourhood(location_text),
            "initial_contact_location":   locs["contact"],
            "incident_location_primary":  locs["incident"],
            # Violence fields intentionally blank — researcher codes these
            "coercion_present": "", "threats_present": "", "physical_force": "",
            "sexual_assault":   "", "robbery_theft":   "", "stealthing":     "", "loss_of_consciousness": "",
            "movement_present": "", "entered_vehicle":  "",
            "suspect_count":              parse_suspect_count(description),
            "suspect_gender":             parse_gender(description),
            "suspect_description_text":   description,
            "suspect_race_ethnicity": "", "suspect_age_estimate": "", "suspect_name": "",
            "vehicle_present":  veh.get("vehicle_present", ""),
            "vehicle_make":     veh.get("vehicle_make",    ""),
            "vehicle_model":    veh.get("vehicle_model",   ""),
            "vehicle_colour":   veh.get("vehicle_colour",  ""),
            "plate_partial":    veh.get("plate_partial",   ""),
            "summary_analytic": "",
            "lat_initial":  lat_initial,
            "lon_initial":  lon_initial,
            "flags": [],
            "_bulletin_text": row_source_text,
        })

    return {"incidents": incidents, "total": len(incidents), "method": "excel"}


from pydantic import BaseModel as PydanticBaseModel

class VisualizeRequest(PydanticBaseModel):
    text: str


@app.post("/nlp/visualize")
def visualize_parse(data: VisualizeRequest):
    """
    Run a text snippet through spaCy and return displaCy SVG/HTML for both
    the dependency parse (dep) and named entity recognition (ent) views.
    Used to verify SVO pattern matching and negation detection.
    """
    from spacy import displacy
    from nlp_analysis import nlp_model

    text = (data.text or "").strip()[:2000]  # truncate — long texts produce unusable SVGs
    if not text:
        raise HTTPException(status_code=400, detail="text must not be empty")
    if nlp_model is None:
        raise HTTPException(status_code=503, detail="spaCy model not loaded")

    doc = nlp_model(text)
    return {
        "dep_html": displacy.render(doc, style="dep", page=False, minify=True),
        "ent_html": displacy.render(doc, style="ent", page=False, minify=True),
    }


class DupCheckItem(PydanticBaseModel):
    index: int
    raw_narrative: str = ""
    incident_date: str = ""
    city: str = ""


class BulkSaveRequest(PydanticBaseModel):
    incidents: list[dict]
    analyst_name: str = ""
    source_organization: str = ""
    bulletin_session_id: str = ""


def _matched_info(report) -> dict:
    """Return a brief preview of a matched DB record for the duplicate review modal."""
    narrative = report.raw_narrative or ""
    return {
        "incident_date": report.incident_date or "",
        "city": report.city or "",
        "narrative_preview": narrative[:120].strip(),
    }


@app.post("/check-duplicates")
def check_duplicates(items: list[DupCheckItem], db: Session = Depends(get_db)):
    """Check a list of parsed incidents against existing reports before import."""

    # Pre-load candidates grouped by date for the fuzzy pass — avoids N×all-reports queries
    all_dates = {item.incident_date.strip() for item in items if item.incident_date.strip()}
    candidates_by_date: dict[str, list] = {}
    for d in all_dates:
        candidates_by_date[d] = db.query(Report).filter(Report.incident_date == d).all()
    # Lazy no-date bucket
    candidates_no_date: list = []

    results = []
    for item in items:
        # 1. Exact match via narrative hash
        if item.raw_narrative.strip():
            h = _narrative_hash(item.raw_narrative)
            exact = db.query(Report).filter(Report.narrative_hash == h).first()
            if exact:
                results.append({
                    "index": item.index, "status": "exact",
                    "matched_report_id": exact.report_id,
                    "matched_info": _matched_info(exact),
                })
                continue

        # 2. Fuzzy narrative match — uses max(Jaccard, overlap-coefficient) so that an
        #    Excel synopsis (short) contained within a full PDF bulletin entry (long)
        #    still scores high even though Jaccard alone would be dragged down by the
        #    extra header/label words in the bulletin.  Threshold 0.45.
        #    Pool: same-date candidates first; fall back to ALL records when pool is
        #    empty (PDF parser may have stored the date differently than Excel parser).
        if item.raw_narrative.strip():
            date = item.incident_date.strip()
            pool = candidates_by_date.get(date, []) if date else []
            if not pool:
                if not candidates_no_date:
                    candidates_no_date.extend(db.query(Report).all())
                pool = candidates_no_date

            best_match = None
            best_score = 0.0
            for c in pool:
                if not c.raw_narrative:
                    continue
                score = _narrative_similarity(item.raw_narrative, c.raw_narrative)
                if score > best_score:
                    best_score = score
                    best_match = c
            if best_match and best_score >= 0.45:
                results.append({
                    "index": item.index, "status": "possible",
                    "matched_report_id": best_match.report_id,
                    "matched_info": _matched_info(best_match),
                })
                continue

        # 3. Possible match: same incident_date AND city (both non-empty)
        date = item.incident_date.strip()
        city = item.city.strip()
        if date and city:
            possible = db.query(Report).filter(
                Report.incident_date == date,
                Report.city.ilike(city),
            ).first()
            if possible:
                results.append({
                    "index": item.index, "status": "possible",
                    "matched_report_id": possible.report_id,
                    "matched_info": _matched_info(possible),
                })
                continue

        results.append({"index": item.index, "status": "new"})

    return {"results": results}


@app.post("/bulk-save")
def bulk_save(data: BulkSaveRequest, db: Session = Depends(get_db)):
    """Save a list of pre-parsed incidents as draft reports."""
    saved = []
    skipped = []
    for inc in data.incidents:
        raw = inc.get("raw_narrative", "")
        if not raw.strip():
            continue

        h = _narrative_hash(raw)
        existing = db.query(Report).filter(Report.narrative_hash == h).first()
        if existing:
            # Back-fill source_bulletin_text if the existing case has none but this import provides it
            bulletin_text = inc.get("_bulletin_text", "")
            if bulletin_text and not (existing.source_bulletin_text or "").strip():
                existing.source_bulletin_text = bulletin_text
                db.commit()
            skipped.append(existing.report_id)
            continue

        report_id = f"RLA-{datetime.utcnow().strftime('%Y%m%d')}-{str(uuid.uuid4())[:6].upper()}"
        report = Report(
            report_id=report_id,
            narrative_hash=h,
            raw_narrative=raw,
            source_bulletin_text=inc.get("_bulletin_text", ""),
            source_bulletin_session_id=inc.get("_session_id", "") or data.bulletin_session_id,
            source_organization=inc.get("source_organization") or data.source_organization,
            analyst_name=data.analyst_name,
            date_received=inc.get("bulletin_date") or datetime.utcnow().strftime("%Y-%m-%d"),
            coding_status="uncoded",
            incident_date=inc.get("incident_date", ""),
            city=inc.get("city", ""),
            neighbourhood=inc.get("neighbourhood", ""),
            initial_contact_location=inc.get("initial_contact_location", ""),
            incident_location_primary=inc.get("incident_location_primary", ""),
            indoor_outdoor=inc.get("indoor_outdoor", ""),
            public_private=inc.get("public_private", ""),
            initial_approach_type=inc.get("initial_approach_type", ""),
            negotiation_present=inc.get("negotiation_present", ""),
            refusal_present=inc.get("refusal_present", ""),
            pressure_after_refusal=inc.get("pressure_after_refusal", ""),
            coercion_present=inc.get("coercion_present", ""),
            threats_present=inc.get("threats_present", ""),
            verbal_abuse=inc.get("verbal_abuse", ""),
            physical_force=inc.get("physical_force", ""),
            sexual_assault=inc.get("sexual_assault", ""),
            robbery_theft=inc.get("robbery_theft", ""),
            stealthing=inc.get("stealthing", ""),
            loss_of_consciousness=inc.get("loss_of_consciousness", ""),
            exit_type=inc.get("exit_type", ""),
            movement_present=inc.get("movement_present", ""),
            movement_attempted=inc.get("movement_attempted", ""),
            entered_vehicle=inc.get("entered_vehicle", ""),
            mode_of_movement=inc.get("mode_of_movement", ""),
            public_to_private_shift=inc.get("public_to_private_shift", ""),
            public_to_secluded_shift=inc.get("public_to_secluded_shift", ""),
            offender_control_over_movement=inc.get("offender_control_over_movement", ""),
            suspect_count=str(inc.get("suspect_count", "")),
            suspect_gender=inc.get("suspect_gender", ""),
            suspect_description_text=inc.get("suspect_description_text", ""),
            suspect_race_ethnicity=inc.get("suspect_race_ethnicity", ""),
            suspect_age_estimate=str(inc.get("suspect_age_estimate", "")),
            vehicle_present=inc.get("vehicle_present", ""),
            vehicle_make=inc.get("vehicle_make", ""),
            vehicle_model=inc.get("vehicle_model", ""),
            vehicle_colour=inc.get("vehicle_colour", ""),
            plate_partial=inc.get("plate_partial", ""),
            summary_analytic=inc.get("summary_analytic", ""),
            lat_initial=inc.get("lat_initial"),
            lon_initial=inc.get("lon_initial"),
            ai_suggestions={"flags": inc.get("flags", []), "entry_type": inc.get("entry_type", "")},
            audit_log=[{"ts": datetime.utcnow().isoformat(), "action": "imported from bulletin", "by": data.analyst_name or "system"}],
        )
        db.add(report)
        saved.append(report_id)

    db.commit()
    return {"saved": len(saved), "report_ids": saved, "skipped": len(skipped), "skipped_report_ids": skipped}


# ── Export ────────────────────────────────────────────────────────────────────

@app.get("/export/csv")
def export_csv(db: Session = Depends(get_db)):
    reports = db.query(Report).all()
    if not reports:
        raise HTTPException(404, "No reports to export")

    output = io.StringIO()
    fieldnames = [c.name for c in Report.__table__.columns if c.name not in ("audit_log", "ai_suggestions")]
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()

    for r in reports:
        row = {f: getattr(r, f, "") for f in fieldnames}
        # Convert lists/dicts to strings for CSV
        for k, v in row.items():
            if isinstance(v, (list, dict)):
                row[k] = json.dumps(v)
            elif v is None:
                row[k] = ""
        writer.writerow(row)

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=redlight_export.csv"},
    )


def _build_geojson_props(r, loc_type: str) -> dict:
    """Shared property dict for a single location feature."""
    _stage_city = {
        "initial_contact": r.initial_contact_city or r.city or "",
        "incident":        r.incident_city or r.city or "",
        "destination":     r.destination_city or r.city or "",
    }
    _stage_city_conf = {
        "initial_contact": r.initial_contact_city_confidence or "",
        "incident":        r.incident_city_confidence or "",
        "destination":     r.destination_city_confidence or "",
    }
    _loc_type_map = {
        "initial_contact": r.initial_contact_location_type or "",
        "incident":        r.incident_location_type or "",
        "destination":     "",
    }
    _geocoding_status_map = {
        "initial_contact": r.initial_contact_geocoding_status or "",
        "incident":        r.incident_geocoding_status or "",
        "destination":     r.destination_geocoding_status or "",
    }
    return {
        # Identity / admin
        "report_id":                       r.report_id,
        "coding_status":                   r.coding_status or "",
        "source_organization":             r.source_organization or "",
        "analyst_name":                    r.analyst_name or "",
        "incident_date":                   r.incident_date or "",
        # Location context
        "location_type":                   loc_type,
        "point_location_type":             _loc_type_map.get(loc_type, ""),
        "geocoding_status":                _geocoding_status_map.get(loc_type, ""),
        "city":                            _stage_city.get(loc_type, r.city or ""),
        "city_confidence":                 _stage_city_conf.get(loc_type, ""),
        "primary_case_city":               r.city or "",
        "neighbourhood":                   r.neighbourhood or "",
        "cross_city_movement":             r.cross_city_movement or "",
        "indoor_outdoor":                  r.indoor_outdoor or "",
        "public_private":                  r.public_private or "",
        "deserted":                        r.deserted or "",
        # Harm indicators
        "primary_incident_type":           r.primary_incident_type or "",
        "overall_severity":                r.overall_severity or "",
        "primary_harm":                    r.primary_harm or "",
        "multi_harm_flag":                 r.multi_harm_flag or "",
        "coercion_present":                r.coercion_present or "",
        "physical_force":                  r.physical_force or "",
        "sexual_assault":                  r.sexual_assault or "",
        "weapon_present_used":             r.weapon_present_used or "",
        "robbery_theft":                   r.robbery_theft or "",
        "restraint_confinement":           r.restraint_confinement or "",
        "prevented_exit":                  r.prevented_exit or "",
        # Mobility
        "movement_present":                r.movement_present or "",
        "mode_of_movement":                r.mode_of_movement or "",
        "entered_vehicle":                 r.entered_vehicle or "",
        "public_to_private_shift":         r.public_to_private_shift or "",
        "offender_control_over_movement":  r.offender_control_over_movement or "",
        # Vehicle
        "vehicle_present":                 r.vehicle_present or "",
        "vehicle_make":                    r.vehicle_make or "",
        "vehicle_colour":                  r.vehicle_colour or "",
        # Exit
        "exit_type":                       r.exit_type or "",
        # Concern flags (analytical — observation only, not confirmed findings)
        "concern_trafficking":             r.concern_trafficking or "",
        "concern_third_party_control":     r.concern_third_party_control or "",
        "concern_grooming":                r.concern_grooming or "",
        "concern_organized_offending":     r.concern_organized_offending or "",
        "concern_repeat_suspect":          r.concern_repeat_suspect or "",
        "concern_repeat_vehicle":          r.concern_repeat_vehicle or "",
        "concern_urgent_public_safety":    r.concern_urgent_public_safety or "",
        "concern_bulletin_suitable":       r.concern_bulletin_suitable or "",
    }


@app.get("/export/geojson")
def export_geojson(
    report_ids: str = "",
    include_lines: bool = False,
    db: Session = Depends(get_db),
):
    """Export cases as GeoJSON point features.

    Optional query params:
    - report_ids: comma-separated IDs to export (omit for all)
    - include_lines: if true, also append LineString features for movement paths
    """
    q = db.query(Report)
    if report_ids:
        id_list = [rid.strip() for rid in report_ids.split(",") if rid.strip()]
        q = q.filter(Report.report_id.in_(id_list))
    reports = q.all()

    features = []
    location_types = [
        ("initial_contact", "lat_initial", "lon_initial"),
        ("incident", "lat_incident", "lon_incident"),
        ("destination", "lat_destination", "lon_destination"),
    ]
    for r in reports:
        for loc_type, lat_col, lon_col in location_types:
            lat = getattr(r, lat_col)
            lon = getattr(r, lon_col)
            if lat is None or lon is None:
                continue
            features.append({
                "type": "Feature",
                "geometry": {"type": "Point", "coordinates": [lon, lat]},
                "properties": _build_geojson_props(r, loc_type),
            })

    if include_lines:
        for r in reports:
            coords = []
            for lat_col, lon_col in [("lat_initial", "lon_initial"), ("lat_incident", "lon_incident"), ("lat_destination", "lon_destination")]:
                lat = getattr(r, lat_col)
                lon = getattr(r, lon_col)
                if lat is not None and lon is not None:
                    coords.append([lon, lat])
            if len(coords) >= 2:
                features.append({
                    "type": "Feature",
                    "geometry": {"type": "LineString", "coordinates": coords},
                    "properties": {
                        "report_id": r.report_id,
                        "location_type": "movement_path",
                        "city": r.city or "",
                        "incident_date": r.incident_date or "",
                        "movement_present": r.movement_present or "",
                        "cross_city_movement": r.cross_city_movement or "",
                        "coding_status": r.coding_status or "",
                    },
                })

    geojson = {"type": "FeatureCollection", "features": features}
    return StreamingResponse(
        iter([json.dumps(geojson, indent=2)]),
        media_type="application/geo+json",
        headers={"Content-Disposition": "attachment; filename=redlight_export.geojson"},
    )


@app.get("/export/shapefile")
def export_shapefile(
    report_ids: str = "",
    include_lines: bool = False,
    db: Session = Depends(get_db),
):
    """Export cases as a zipped Shapefile bundle (QGIS-ready).

    Optional query params:
    - report_ids: comma-separated IDs to export (omit for all)
    - include_lines: if true, also include a movement LineString shapefile
    """
    import shapefile
    import zipfile

    q = db.query(Report)
    if report_ids:
        id_list = [rid.strip() for rid in report_ids.split(",") if rid.strip()]
        q = q.filter(Report.report_id.in_(id_list))
    reports = q.all()

    location_types = [
        ("initial_contact", "lat_initial", "lon_initial"),
        ("incident", "lat_incident", "lon_incident"),
        ("destination", "lat_destination", "lon_destination"),
    ]

    # ── Point shapefile ────────────────────────────────────────────────────────
    pt_shp = io.BytesIO()
    pt_shx = io.BytesIO()
    pt_dbf = io.BytesIO()
    with shapefile.Writer(shp=pt_shp, shx=pt_shx, dbf=pt_dbf, shapeType=shapefile.POINT) as w:
        w.field("report_id",  "C", 64)
        w.field("loc_type",   "C", 32)
        w.field("city",       "C", 64)
        w.field("inc_date",   "C", 20)
        w.field("coercion",   "C", 16)
        w.field("movement",   "C", 16)
        w.field("phys_force", "C", 16)
        w.field("sex_aslt",   "C", 16)
        w.field("veh_prsnt",  "C", 16)
        w.field("exit_type",  "C", 32)
        w.field("pub_priv",   "C", 16)
        w.field("offndr_ctrl","C", 32)
        w.field("status",     "C", 32)
        w.field("source_org", "C", 64)
        w.field("cross_city", "C", 16)
        w.field("nbhd",       "C", 64)
        for r in reports:
            for loc_type, lat_col, lon_col in location_types:
                lat = getattr(r, lat_col)
                lon = getattr(r, lon_col)
                if lat is None or lon is None:
                    continue
                w.point(lon, lat)
                w.record(
                    r.report_id or "",
                    loc_type,
                    r.city or "",
                    r.incident_date or "",
                    r.coercion_present or "",
                    r.movement_present or "",
                    r.physical_force or "",
                    r.sexual_assault or "",
                    r.vehicle_present or "",
                    r.exit_type or "",
                    r.public_to_private_shift or "",
                    r.offender_control_over_movement or "",
                    r.coding_status or "",
                    r.source_organization or "",
                    r.cross_city_movement or "",
                    r.neighbourhood or "",
                )

    # ── Movement LineString shapefile (optional) ───────────────────────────────
    ln_shp = ln_shx = ln_dbf = None
    if include_lines:
        ln_shp = io.BytesIO()
        ln_shx = io.BytesIO()
        ln_dbf = io.BytesIO()
        with shapefile.Writer(shp=ln_shp, shx=ln_shx, dbf=ln_dbf, shapeType=shapefile.POLYLINE) as w:
            w.field("report_id",  "C", 64)
            w.field("city",       "C", 64)
            w.field("inc_date",   "C", 20)
            w.field("movement",   "C", 16)
            w.field("cross_city", "C", 16)
            w.field("status",     "C", 32)
            for r in reports:
                coords = []
                for lat_col, lon_col in [("lat_initial", "lon_initial"), ("lat_incident", "lon_incident"), ("lat_destination", "lon_destination")]:
                    lat = getattr(r, lat_col)
                    lon = getattr(r, lon_col)
                    if lat is not None and lon is not None:
                        coords.append([lon, lat])
                if len(coords) >= 2:
                    w.line([coords])
                    w.record(
                        r.report_id or "",
                        r.city or "",
                        r.incident_date or "",
                        r.movement_present or "",
                        r.cross_city_movement or "",
                        r.coding_status or "",
                    )

    # ── Bundle into ZIP ────────────────────────────────────────────────────────
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("redlight_points.shp", pt_shp.getvalue())
        zf.writestr("redlight_points.shx", pt_shx.getvalue())
        zf.writestr("redlight_points.dbf", pt_dbf.getvalue())
        if include_lines and ln_shp:
            zf.writestr("redlight_movements.shp", ln_shp.getvalue())
            zf.writestr("redlight_movements.shx", ln_shx.getvalue())
            zf.writestr("redlight_movements.dbf", ln_dbf.getvalue())

    zip_buf.seek(0)
    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=redlight_shapefile.zip"},
    )


# ── Similarity / linkage ─────────────────────────────────────────────────────

@app.get("/reports/{report_id}/similar")
def get_similar_cases(
    report_id: str,
    limit: int = 20,
    min_score: float = 10.0,
    db: Session = Depends(get_db),
):
    target = db.query(Report).filter(Report.report_id == report_id).first()
    if not target:
        raise HTTPException(404, "Report not found")

    others = db.query(Report).filter(Report.report_id != report_id).all()
    results = []
    for r in others:
        sim = compute_similarity(target, r)
        if sim['score'] >= min_score:
            results.append({
                'report_id': r.report_id,
                'incident_date': r.incident_date or '',
                'city': r.city or '',
                'neighbourhood': r.neighbourhood or '',
                'coding_status': r.coding_status or '',
                'coercion_present': r.coercion_present or '',
                'movement_present': r.movement_present or '',
                'vehicle_present': r.vehicle_present or '',
                'physical_force': r.physical_force or '',
                'sexual_assault': r.sexual_assault or '',
                'suspect_gender': r.suspect_gender or '',
                'vehicle_make': r.vehicle_make or '',
                'vehicle_colour': r.vehicle_colour or '',
                'plate_partial': r.plate_partial or '',
                'similarity': sim,
                # existing linkage status if any
                'linkage_status': '',
            })

    results.sort(key=lambda x: x['similarity']['score'], reverse=True)

    # Attach existing linkage assessments
    ids = [x['report_id'] for x in results[:limit]]
    linkages = db.query(CaseLinkage).filter(
        ((CaseLinkage.report_id_a == report_id) | (CaseLinkage.report_id_b == report_id))
    ).all()
    linkage_map = {}
    for lk in linkages:
        other = lk.report_id_b if lk.report_id_a == report_id else lk.report_id_a
        linkage_map[other] = lk.analyst_status

    for item in results[:limit]:
        item['linkage_status'] = linkage_map.get(item['report_id'], '')

    return results[:limit]


@app.get("/reports/{report_id_a}/compare/{report_id_b}")
def compare_reports(report_id_a: str, report_id_b: str, db: Session = Depends(get_db)):
    a = db.query(Report).filter(Report.report_id == report_id_a).first()
    b = db.query(Report).filter(Report.report_id == report_id_b).first()
    if not a or not b:
        raise HTTPException(404, "One or both reports not found")

    sim = compute_similarity(a, b)

    linkage = db.query(CaseLinkage).filter(
        ((CaseLinkage.report_id_a == report_id_a) & (CaseLinkage.report_id_b == report_id_b)) |
        ((CaseLinkage.report_id_a == report_id_b) & (CaseLinkage.report_id_b == report_id_a))
    ).first()

    return {
        'report_a': ReportOut.model_validate(a).model_dump(),
        'report_b': ReportOut.model_validate(b).model_dump(),
        'similarity': sim,
        'linkage': {
            'analyst_status': linkage.analyst_status,
            'analyst_notes': linkage.analyst_notes,
        } if linkage else None,
    }


class LinkageUpdate(PydanticBaseModel):
    report_id_a: str
    report_id_b: str
    analyst_status: str
    analyst_notes: str = ""


@app.post("/linkage")
def save_linkage(data: LinkageUpdate, db: Session = Depends(get_db)):
    existing = db.query(CaseLinkage).filter(
        ((CaseLinkage.report_id_a == data.report_id_a) & (CaseLinkage.report_id_b == data.report_id_b)) |
        ((CaseLinkage.report_id_a == data.report_id_b) & (CaseLinkage.report_id_b == data.report_id_a))
    ).first()

    if existing:
        existing.analyst_status = data.analyst_status
        existing.analyst_notes = data.analyst_notes
        existing.updated_at = datetime.utcnow()
    else:
        # Compute score to store
        ra = db.query(Report).filter(Report.report_id == data.report_id_a).first()
        rb = db.query(Report).filter(Report.report_id == data.report_id_b).first()
        sim = compute_similarity(ra, rb) if ra and rb else {}
        lk = CaseLinkage(
            report_id_a=data.report_id_a,
            report_id_b=data.report_id_b,
            similarity_score=sim.get('score', 0.0),
            score_breakdown=sim.get('dimensions', {}),
            analyst_status=data.analyst_status,
            analyst_notes=data.analyst_notes,
        )
        db.add(lk)

    db.commit()
    return {"ok": True}


# ── Stats / patterns ─────────────────────────────────────────────────────────

@app.get("/stats")
def get_stats(db: Session = Depends(get_db)):
    from nlp_analysis import _nlp as _spacy_model
    reports = db.query(Report).all()
    total = len(reports)

    def pct(field, val):
        count = sum(1 for r in reports if getattr(r, field) == val)
        return {"count": count, "pct": round(count / total * 100, 1) if total else 0}

    # Repeated vehicles
    vehicles = [r.plate_partial for r in reports if r.plate_partial]
    from collections import Counter
    vehicle_counts = Counter(vehicles)
    repeated_vehicles = [{"plate": p, "count": c} for p, c in vehicle_counts.most_common(10) if c > 1]

    # Vehicle makes — normalise to title-case for deduplication
    makes = [r.vehicle_make.strip().title() for r in reports if r.vehicle_make and r.vehicle_make.strip()]
    make_counts = Counter(makes)

    neighbourhoods = [r.neighbourhood.strip().title() for r in reports if r.neighbourhood and r.neighbourhood.strip()]
    neighbourhood_counts = Counter(neighbourhoods)

    _SKIP_CITIES = {'', 'none', 'unknown', 'n/a', 'probable', 'inferred'}
    # Collect from all stage-specific city fields; fall back to legacy summary city.
    # For each report, prefer stage-specific over summary; de-duplicate within report.
    city_entries = []
    for r in reports:
        stage_cities = {
            c.strip().title()
            for c in [r.initial_contact_city or '', r.incident_city or '', r.destination_city or '']
            if c and c.strip().lower() not in _SKIP_CITIES
        }
        if stage_cities:
            city_entries.extend(stage_cities)
        elif r.city and r.city.strip().lower() not in _SKIP_CITIES:
            city_entries.append(r.city.strip().title())
    city_counts = Counter(city_entries)

    # Vehicle colours — normalise case
    colours = [r.vehicle_colour.strip().title() for r in reports if r.vehicle_colour and r.vehicle_colour.strip()]
    colour_counts = Counter(colours)

    # Vehicle types (stored in vehicle_model after import) — normalise
    _KNOWN_TYPES = {"sedan", "suv", "van", "minivan", "truck", "pickup", "hatchback",
                    "coupe", "convertible", "wagon", "cab", "taxi"}
    vtypes = [r.vehicle_model.strip().title() for r in reports
              if r.vehicle_model and r.vehicle_model.strip().lower() in _KNOWN_TYPES]
    vtype_counts = Counter(vtypes)

    # Approach type
    foot_count = sum(1 for r in reports if r.mode_of_movement == "foot")
    vehicle_approach = sum(1 for r in reports if r.mode_of_movement == "vehicle")

    # Year breakdown
    year_counts: Counter = Counter()
    for r in reports:
        if r.incident_date and len(r.incident_date) >= 4:
            try:
                year = int(r.incident_date[:4])
                if 2000 <= year <= 2100:
                    year_counts[year] += 1
            except ValueError:
                pass

    # ── NLP violence detection counts (from ai_suggestions, pre-coding) ──────
    def _nlp_rank(field: str, rank: int) -> int:
        return sum(
            1 for r in reports
            if (r.ai_suggestions or {}).get("nlp", {}).get(field) == rank
        )

    def _nlp_esc(min_score: int) -> int:
        return sum(
            1 for r in reports
            if (r.ai_suggestions or {}).get("nlp", {}).get("escalation", {}).get("score", 0) >= min_score
        )

    nlp_violence = {
        "coercion":  {"rank1": _nlp_rank("coercion_rank", 1),  "rank2": _nlp_rank("coercion_rank", 2)},
        "physical":  {"rank1": _nlp_rank("physical_rank", 1),  "rank2": _nlp_rank("physical_rank", 2)},
        "sexual":    {"rank1": _nlp_rank("sexual_rank", 1),    "rank2": _nlp_rank("sexual_rank", 2)},
        "movement":  {"rank1": _nlp_rank("movement_rank", 1),  "rank2": _nlp_rank("movement_rank", 2)},
        "weapon":    {"rank1": _nlp_rank("weapon_rank", 1),    "rank2": _nlp_rank("weapon_rank", 2)},
        "escalation": {
            "score3": _nlp_esc(3),
            "score4": _nlp_esc(4),
            "score5": _nlp_esc(5),
        },
    }

    # Named escalation patterns across all cases
    pattern_counter: Counter = Counter()
    for r in reports:
        patterns = (r.ai_suggestions or {}).get("nlp", {}).get("escalation", {}).get("patterns", [])
        for p in patterns:
            pattern_counter[p] += 1

    return {
        "total": total,
        "coded": sum(1 for r in reports if r.coding_status == "coded"),
        "coercion": pct("coercion_present", "yes"),
        "movement": pct("movement_present", "yes"),
        "physical_force": pct("physical_force", "yes"),
        "sexual_assault": pct("sexual_assault", "yes"),
        "threats_present": pct("threats_present", "yes"),
        "vehicle_present": pct("vehicle_present", "yes"),
        "vehicle_present_count": sum(1 for r in reports if r.vehicle_present == "yes") or sum(1 for r in reports if r.mode_of_movement == "vehicle"),
        "nlp_violence": nlp_violence,
        "nlp_escalation_patterns": [{"pattern": p, "count": c} for p, c in pattern_counter.most_common(10)],
        "repeated_vehicles": repeated_vehicles,
        "vehicle_makes": [{"make": m, "count": c} for m, c in make_counts.most_common(10)],
        "vehicle_colours": [{"colour": c, "count": n} for c, n in colour_counts.most_common(8)],
        "vehicle_types": [{"type": t, "count": c} for t, c in vtype_counts.most_common(8)],
        "approach_foot": foot_count,
        "approach_vehicle": vehicle_approach,
        "year_breakdown": [{"year": y, "count": c} for y, c in sorted(year_counts.items())],
        "neighbourhoods": [{"name": n, "count": c} for n, c in neighbourhood_counts.most_common(10)],
        "cities": [{"name": c, "count": n} for c, n in city_counts.most_common(15)],
        "nlp_available": _spacy_model is not None,
        "map_points": [
            {
                "report_id": r.report_id,
                "lat_initial": r.lat_initial,
                "lon_initial": r.lon_initial,
                "lat_incident": r.lat_incident,
                "lon_incident": r.lon_incident,
                "lat_destination": r.lat_destination,
                "lon_destination": r.lon_destination,
                # harm flags
                "coercion": r.coercion_present,
                "physical_force": r.physical_force,
                "sexual_assault": r.sexual_assault,
                "robbery_theft": r.robbery_theft,
                # movement
                "movement": r.movement_present,
                "movement_completed": r.movement_completed,
                "entered_vehicle": r.entered_vehicle,
                "mode_of_movement": r.mode_of_movement or '',
                "offender_control_over_movement": r.offender_control_over_movement or '',
                "public_to_private_shift": r.public_to_private_shift,
                "public_to_secluded_shift": r.public_to_secluded_shift or '',
                "cross_municipality": r.cross_municipality,
                # sequence
                "highest_stage_reached": r.highest_stage_reached,
                # meta
                "city": r.city,
                "incident_date": r.incident_date,
                "coding_status": r.coding_status,
                # location confidence
                "location_certainty": r.location_certainty,
                "initial_contact_city_confidence": r.initial_contact_city_confidence,
                "incident_city_confidence": r.incident_city_confidence,
                "destination_city_confidence": r.destination_city_confidence,
                # enriched fields
                "neighbourhood":                    r.neighbourhood or '',
                "repeat_suspect_flag":              r.repeat_suspect_flag or '',
                "known_repeat_suspect":             r.known_repeat_suspect or '',
                "repeat_vehicle_flag":              r.repeat_vehicle_flag or '',
                "vehicle_present":                  r.vehicle_present or '',
                "plate_partial":                    r.plate_partial or '',
                "suspect_distinctive_features":     r.suspect_distinctive_features or '',
                "vehicle_description":              r.vehicle_description or '',
                "initial_contact_address_raw":      r.initial_contact_address_raw or '',
                "initial_contact_precision":        r.initial_contact_precision or '',
                "initial_contact_geocoding_status": r.initial_contact_geocoding_status or '',
                "incident_address_raw":             r.incident_address_raw or '',
                "incident_precision":               r.incident_precision or '',
                "incident_geocoding_status":        r.incident_geocoding_status or '',
                "primary_harm":                     r.primary_harm or '',
            }
            for r in reports
            if r.lat_initial or r.lat_incident
        ],
    }


# ── Stage CRUD ────────────────────────────────────────────────────────────────

@app.get("/reports/{report_id}/stages", response_model=list[StageOut])
def list_stages(report_id: str, db: Session = Depends(get_db)):
    return (
        db.query(ReportStage)
        .filter(ReportStage.report_id == report_id)
        .order_by(ReportStage.stage_order)
        .all()
    )


@app.post("/reports/{report_id}/stages", response_model=StageOut)
def create_stage(report_id: str, body: StageCreate, db: Session = Depends(get_db)):
    stage = ReportStage(
        report_id=report_id,
        stage_order=body.stage_order or 1,
        stage_type=body.stage_type or "",
        client_behaviors=body.client_behaviors or [],
        victim_responses=body.victim_responses or [],
        turning_point_notes=body.turning_point_notes or "",
        visibility=body.visibility or "",
        guardianship=body.guardianship or "",
        isolation_level=body.isolation_level or "",
        control_type=body.control_type or "",
        location_label=body.location_label or "",
        location_type=body.location_type or "",
        movement_type_to_here=body.movement_type_to_here or "",
    )
    db.add(stage)
    db.commit()
    db.refresh(stage)
    return stage


@app.put("/reports/{report_id}/stages/reorder")
def reorder_stages(report_id: str, items: list[StageReorderItem], db: Session = Depends(get_db)):
    for item in items:
        stage = db.query(ReportStage).filter(
            ReportStage.id == item.id,
            ReportStage.report_id == report_id,
        ).first()
        if stage:
            stage.stage_order = item.stage_order
    db.commit()
    return {"ok": True}


@app.put("/reports/{report_id}/stages/{stage_id}", response_model=StageOut)
def update_stage(report_id: str, stage_id: int, body: StageUpdate, db: Session = Depends(get_db)):
    stage = db.query(ReportStage).filter(
        ReportStage.id == stage_id,
        ReportStage.report_id == report_id,
    ).first()
    if not stage:
        raise HTTPException(404, "Stage not found")
    for field, value in body.model_dump(exclude_none=True).items():
        setattr(stage, field, value)
    db.commit()
    db.refresh(stage)
    return stage


@app.delete("/reports/{report_id}/stages/{stage_id}")
def delete_stage(report_id: str, stage_id: int, db: Session = Depends(get_db)):
    stage = db.query(ReportStage).filter(
        ReportStage.id == stage_id,
        ReportStage.report_id == report_id,
    ).first()
    if not stage:
        raise HTTPException(404, "Stage not found")
    db.delete(stage)
    db.commit()
    return {"ok": True}


# ── Research: stage patterns ───────────────────────────────────────────────────

@app.get("/research/stage-patterns")
def get_stage_patterns(
    stage_type:       Optional[str] = None,
    visibility:       Optional[str] = None,
    guardianship:     Optional[str] = None,
    isolation:        Optional[str] = None,
    escalation_level: Optional[str] = None,
    date_from:        Optional[str] = None,
    date_to:          Optional[str] = None,
    db: Session = Depends(get_db),
):
    """
    Cross-case stage analysis.
    Returns:
      - stage_type_frequency: count per stage type
      - visibility_by_stage: distribution of visibility values per stage type
      - guardianship_by_stage: distribution of guardianship per stage type
      - isolation_by_stage: distribution of isolation per stage type
      - control_by_stage: distribution of control type per stage type
      - behavior_frequency: client_behavior code counts across all stages
      - response_frequency: victim_response code counts across all stages
      - movement_by_stage: movement_type_to_here distribution per stage type
      - matching_cases: report_ids of cases with any stage matching the filter params
    """
    from collections import Counter, defaultdict

    query = db.query(ReportStage)
    if stage_type:
        query = query.filter(ReportStage.stage_type == stage_type)
    if visibility:
        query = query.filter(ReportStage.visibility == visibility)
    if guardianship:
        query = query.filter(ReportStage.guardianship == guardianship)
    if isolation:
        query = query.filter(ReportStage.isolation_level == isolation)
    if escalation_level:
        query = query.filter(ReportStage.escalation_level == escalation_level)
    if date_from or date_to:
        # Join to Report to filter by incident_date
        query = query.join(Report, ReportStage.report_id == Report.report_id)
        if date_from:
            query = query.filter(Report.incident_date >= date_from)
        if date_to:
            query = query.filter(Report.incident_date <= date_to)

    stages = query.all()

    # Frequency counters
    type_freq: Counter = Counter()
    vis_by_stage: dict = defaultdict(Counter)
    guard_by_stage: dict = defaultdict(Counter)
    iso_by_stage: dict = defaultdict(Counter)
    ctrl_by_stage: dict = defaultdict(Counter)
    move_by_stage: dict = defaultdict(Counter)
    behavior_freq: Counter = Counter()
    response_freq: Counter = Counter()
    escalation_freq: Counter = Counter()
    movement_impact_freq: Counter = Counter()
    matching_cases: set = set()

    for s in stages:
        t = s.stage_type or "unknown"
        type_freq[t] += 1
        if s.visibility:   vis_by_stage[t][s.visibility] += 1
        if s.guardianship: guard_by_stage[t][s.guardianship] += 1
        if s.isolation_level: iso_by_stage[t][s.isolation_level] += 1
        if s.control_type:    ctrl_by_stage[t][s.control_type] += 1
        if s.movement_type_to_here: move_by_stage[t][s.movement_type_to_here] += 1
        for b in (s.client_behaviors or []):
            behavior_freq[b] += 1
        for r in (s.victim_responses or []):
            response_freq[r] += 1
        if getattr(s, 'escalation_level', None):
            escalation_freq[s.escalation_level] += 1
        if getattr(s, 'movement_impact', None):
            movement_impact_freq[s.movement_impact] += 1
        matching_cases.add(s.report_id)

    def _counter_to_list(c: Counter):
        return [{"value": k, "count": v} for k, v in c.most_common()]

    def _nested_to_dict(d: dict):
        return {k: _counter_to_list(v) for k, v in d.items()}

    # Per-case stage sequences (for cross-case grouping)
    all_stages = db.query(ReportStage).order_by(ReportStage.report_id, ReportStage.stage_order).all()
    seq_map: dict = defaultdict(list)
    for s in all_stages:
        seq_map[s.report_id].append(s.stage_type or "?")
    seq_strings: Counter = Counter(" → ".join(v) for v in seq_map.values() if v)

    return {
        "stage_type_frequency":      _counter_to_list(type_freq),
        "visibility_by_stage":       _nested_to_dict(vis_by_stage),
        "guardianship_by_stage":     _nested_to_dict(guard_by_stage),
        "isolation_by_stage":        _nested_to_dict(iso_by_stage),
        "control_by_stage":          _nested_to_dict(ctrl_by_stage),
        "movement_by_stage":         _nested_to_dict(move_by_stage),
        "behavior_frequency":        _counter_to_list(behavior_freq),
        "response_frequency":        _counter_to_list(response_freq),
        "escalation_frequency":      _counter_to_list(escalation_freq),
        "movement_impact_frequency": _counter_to_list(movement_impact_freq),
        "matching_cases":            sorted(matching_cases),
        "sequence_frequency":        _counter_to_list(seq_strings),
        "total_stages":              len(stages),
        "total_cases_with_stages":   len(seq_map),
    }


# ── Research analysis ─────────────────────────────────────────────────────────

@app.get("/research/aggregate")
def get_research_aggregate(db: Session = Depends(get_db)):
    """Full cross-case research analysis: sequences, mobility, environment, encounter overview."""
    from research import aggregate_sequences, aggregate_mobility, aggregate_environment, aggregate_encounter, aggregate_vawg
    from sqlalchemy import func as sqlfunc
    reports = db.query(Report).all()

    def _coded(r, field: str) -> bool:
        """True if field has a non-empty analyst-coded value."""
        fp = (getattr(r, 'field_provenance', None) or {})
        if isinstance(fp, dict):
            prov = fp.get(field, 'unset')
        else:
            prov = 'unset'
        val = (getattr(r, field, None) or '').strip()
        return bool(val) and prov in ('analyst_filled', 'reviewed')

    from models import ReportStage
    stage_report_ids = {s.report_id for s in db.query(ReportStage.report_id).distinct().all()}

    _POSITIVE_VALS = {'yes', 'probable', 'inferred', 'probable / inferred'}
    _HARM_FIELDS = [
        'coercion_present', 'threats_present', 'physical_force', 'sexual_assault',
        'stealthing', 'robbery_theft', 'forced_movement_dragging', 'restraint_confinement',
        'weapon_present_used', 'choking_strangulation', 'prevented_exit',
        'non_consensual_substance', 'loss_of_consciousness',
    ]
    _VAWG_FIELDS = [
        'trafficking_exploitation_concern', 'third_party_control_indicated',
        'grooming_recruitment_concern', 'organized_group_offending_concern',
        'worker_appears_controlled', 'client_connected_to_controller',
        'movement_to_unknown_unsafe_location', 'worker_unaware_how_arrived',
        'repeat_targeting_concern', 'multiple_women_referenced',
    ]

    # Data quality summary
    dq = {
        'total_imported':            len(reports),
        'with_encounter_coded':      sum(1 for r in reports if _coded(r, 'primary_incident_type') or _coded(r, 'overall_severity')),
        'with_stage_coding':         len(stage_report_ids),
        'with_location_coded':       sum(1 for r in reports if (r.initial_contact_location or r.incident_location_primary or '').strip()),
        'with_movement_coded':       sum(1 for r in reports if (r.movement_present or '').strip()),
        'with_vawg_coded':           sum(1 for r in reports if any(
            (getattr(r, f, '') or '').strip().lower() in _POSITIVE_VALS
            for f in _VAWG_FIELDS
        )),
        'with_severity_coded':       sum(1 for r in reports if (r.overall_severity or '').strip()),
        'with_suitability_coded':    sum(1 for r in reports if (r.stage_coding_suitability or '').strip()),
        'with_clarity_coded':        sum(1 for r in reports if (r.sequence_clarity or '').strip()),
        'with_harm_coded':           sum(1 for r in reports if any(
            (getattr(r, f, '') or '').strip().lower() in _POSITIVE_VALS
            for f in _HARM_FIELDS
        )),
    }

    def _val_counts(field: str) -> dict:
        from collections import Counter
        c: Counter = Counter()
        for r in reports:
            val = (getattr(r, field, None) or '').strip()
            if val:
                c[val] += 1
        return {'counts': dict(c), 'total_coded': sum(c.values())}

    codability = {
        'narrative_detail_level':      _val_counts('narrative_detail_level'),
        'sequence_reconstructable':    _val_counts('sequence_reconstructable'),
        'stage_coding_suitability':    _val_counts('stage_coding_suitability'),
        'movement_coding_suitability': _val_counts('movement_coding_suitability'),
        'location_coding_suitability': _val_counts('location_coding_suitability'),
        'main_data_limitation':        _val_counts('main_data_limitation'),
        'initial_contact_visible':     _val_counts('initial_contact_visible'),
        'negotiation_visible':         _val_counts('negotiation_visible'),
        'movement_visible':            _val_counts('movement_visible'),
        'violence_coercion_visible':   _val_counts('violence_coercion_visible'),
        'exit_aftermath_visible':      _val_counts('exit_aftermath_visible'),
        'movement_pattern_type':       _val_counts('movement_pattern_type'),
        'movement_timing':             _val_counts('movement_timing'),
        'mappable_status':             _val_counts('mappable_status'),
        'primary_setting_type':        _val_counts('primary_setting_type'),
        'visibility_case':             _val_counts('visibility_case'),
        'isolation_case':              _val_counts('isolation_case'),
        'guardianship_case':           _val_counts('guardianship_case'),
        'access_to_help':              _val_counts('access_to_help'),
        'setting_control':             _val_counts('setting_control'),
        'sequence_pattern':            _val_counts('sequence_pattern'),
        'highest_stage_reached':       _val_counts('highest_stage_reached'),
    }

    return {
        'sequences':    aggregate_sequences(reports),
        'mobility':     aggregate_mobility(reports),
        'environment':  aggregate_environment(reports),
        'encounter':    aggregate_encounter(reports),
        'vawg':         aggregate_vawg(reports),
        'data_quality': dq,
        'codability':   codability,
        'total':        len(reports),
    }


@app.get("/research/linkage-patterns")
def get_linkage_patterns(db: Session = Depends(get_db)):
    """
    Aggregate potential linkage signals across all reports:
    - repeated_vehicles: plates / make+colour combos seen in 2+ cases
    - repeated_locations: initial contact / incident locations in 2+ cases
    - behavior_clusters: co-occurring violence indicator patterns in 2+ cases
    """
    from collections import Counter, defaultdict

    reports = db.query(Report).all()

    # ── Repeated plates ───────────────────────────────────────────────────────
    plate_map: dict = defaultdict(list)
    for r in reports:
        p = (r.plate_partial or "").strip().upper()
        if p:
            plate_map[p].append(r.report_id)

    repeated_vehicles = []
    # Plates with 2+ cases
    for plate, rids in plate_map.items():
        if len(rids) >= 2:
            repeated_vehicles.append({"descriptor": plate, "count": len(rids), "report_ids": rids, "type": "plate"})

    # Make + colour combos with 2+ cases
    make_colour_map: dict = defaultdict(list)
    for r in reports:
        make = (r.vehicle_make or "").strip().title()
        colour = (r.vehicle_colour or "").strip().title()
        if make and colour:
            key = f"{colour} {make}"
            make_colour_map[key].append(r.report_id)
    for desc, rids in make_colour_map.items():
        if len(rids) >= 2:
            repeated_vehicles.append({"descriptor": desc, "count": len(rids), "report_ids": rids, "type": "make_colour"})

    repeated_vehicles.sort(key=lambda x: -x["count"])

    # ── Repeated locations ────────────────────────────────────────────────────
    # Normalisation helpers
    _NARRATIVE_STARTS = (
        'the ', 'it ', 'she ', 'he ', 'they ', 'worker ', 'suspect ',
        'client ', 'victim ', 'this ', 'an unknown', 'not known',
        'communicated', 'contact was', 'foot ',
    )
    _NON_GEO = ('text', 'phone', 'email', 'online', 'app', 'communicated',
                'unknown where', 'unknown pickup', 'unknown location', 'unclear',
                'not specified', 'not known')

    def _is_usable_location(raw: str) -> bool:
        """Return True only if raw looks like a geographic descriptor."""
        s = raw.strip()
        if len(s) < 4 or len(s) > 80:
            return False
        lc = s.lower()
        if any(lc.startswith(p) for p in _NARRATIVE_STARTS):
            return False
        if any(p in lc for p in _NON_GEO):
            return False
        if '.' in s and len(s) > 20:   # sentence with a full-stop
            return False
        return True

    def _normalize_location(raw: str) -> str:
        """
        Canonical form for a location string so near-duplicates merge.
        - title-case
        - sort the two halves of an intersection so order doesn't matter
          e.g. "Windsor and Kingsway" == "Kingsway and Windsor"
        """
        s = raw.strip().title()
        # Normalise connector variations
        for sep in (' And ', ' & ', '/', '\\'):
            if sep in s:
                parts = [p.strip() for p in s.split(sep, 1)]
                parts.sort()
                return f"{parts[0]} and {parts[1]}"
        return s

    loc_map: dict = defaultdict(list)
    for r in reports:
        for loc_field in ["initial_contact_location", "incident_location_primary"]:
            loc = (getattr(r, loc_field, None) or "").strip()
            if not loc or not _is_usable_location(loc):
                continue
            key = _normalize_location(loc)
            if r.report_id not in loc_map[key]:
                loc_map[key].append(r.report_id)

    repeated_locations = []
    for loc, rids in loc_map.items():
        if len(rids) >= 2:
            repeated_locations.append({"descriptor": loc, "count": len(rids), "report_ids": rids})
    repeated_locations.sort(key=lambda x: -x["count"])

    # ── Behavior clusters ─────────────────────────────────────────────────────
    _FLAGS = [
        ("coercion_present",  "Coercion"),
        ("threats_present",   "Threats"),
        ("physical_force",    "Physical force"),
        ("sexual_assault",    "Sexual assault"),
        ("movement_present",  "Movement"),
        ("entered_vehicle",   "Vehicle entry"),
    ]
    cluster_map: dict = defaultdict(list)
    for r in reports:
        active = [label for field, label in _FLAGS if getattr(r, field, "") == "yes"]
        if len(active) >= 2:
            key = " + ".join(active)
            cluster_map[key].append(r.report_id)

    behavior_clusters = [
        {"descriptor": k, "count": len(v), "report_ids": v}
        for k, v in cluster_map.items()
        if len(v) >= 2
    ]
    behavior_clusters.sort(key=lambda x: -x["count"])

    return {
        "repeated_vehicles":  repeated_vehicles[:20],
        "repeated_locations": repeated_locations[:20],
        "behavior_clusters":  behavior_clusters[:20],
    }


# ── Research Notes CRUD ───────────────────────────────────────────────────────

class ResearchNoteCreate(BaseModel):
    note_text: str
    tagged_report_ids: list = []
    tagged_pattern: str = ""


@app.get("/research/notes")
def list_research_notes(db: Session = Depends(get_db)):
    notes = db.query(ResearchNote).order_by(ResearchNote.created_at.desc()).all()
    return [
        {
            "id": n.id,
            "note_text": n.note_text,
            "tagged_report_ids": n.tagged_report_ids or [],
            "tagged_pattern": n.tagged_pattern or "",
            "created_at": n.created_at.isoformat() if n.created_at else "",
        }
        for n in notes
    ]


@app.post("/research/notes")
def create_research_note(body: ResearchNoteCreate, db: Session = Depends(get_db)):
    note = ResearchNote(
        note_text=body.note_text,
        tagged_report_ids=body.tagged_report_ids,
        tagged_pattern=body.tagged_pattern,
        created_at=datetime.utcnow(),
    )
    db.add(note)
    db.commit()
    db.refresh(note)
    return {
        "id": note.id,
        "note_text": note.note_text,
        "tagged_report_ids": note.tagged_report_ids or [],
        "tagged_pattern": note.tagged_pattern or "",
        "created_at": note.created_at.isoformat() if note.created_at else "",
    }


@app.delete("/research/notes/{note_id}")
def delete_research_note(note_id: int, db: Session = Depends(get_db)):
    note = db.query(ResearchNote).filter(ResearchNote.id == note_id).first()
    if not note:
        raise HTTPException(404, "Note not found")
    db.delete(note)
    db.commit()
    return {"ok": True}


# ── Bulletin data export ──────────────────────────────────────────────────────

@app.get("/export/bulletin-data")
def get_bulletin_data(
    date_from: Optional[str] = None,
    date_to:   Optional[str] = None,
    status:    Optional[str] = None,
    city:      Optional[str] = None,
    db: Session = Depends(get_db),
):
    """
    Return all data needed to render a structured analytic bulletin.
    Filters: date_from, date_to, status (coding_status), city.
    Sections: meta, overview, map_points, behavioral, conditions, movement, linkage.
    """
    from collections import Counter
    from research import aggregate_sequences, aggregate_mobility, aggregate_environment

    query = db.query(Report)
    if date_from:
        query = query.filter(Report.incident_date >= date_from)
    if date_to:
        query = query.filter(Report.incident_date <= date_to)
    if status:
        query = query.filter(Report.coding_status == status)
    if city:
        query = query.filter(
            (Report.city.ilike(f"%{city}%")) |
            (Report.initial_contact_city.ilike(f"%{city}%")) |
            (Report.incident_city.ilike(f"%{city}%"))
        )

    reports = query.all()
    total = len(reports)

    if total == 0:
        return {
            "meta": {"case_count": 0, "date_from": date_from, "date_to": date_to, "status": status, "city": city},
            "overview": {}, "map_points": [], "behavioral": {}, "conditions": {}, "movement": {}, "linkage": {},
        }

    # ── Sections ──────────────────────────────────────────────────────────────
    dates = [r.incident_date for r in reports if r.incident_date]
    cities_ctr: Counter = Counter()
    for r in reports:
        for c in [r.city, r.initial_contact_city, r.incident_city]:
            if c and c.strip():
                cities_ctr[c.strip().title()] += 1

    location_types = Counter()
    for r in reports:
        lt = r.destination_location_type or r.start_location_type
        if lt:
            location_types[lt] += 1

    geocoded_count = sum(
        1 for r in reports
        if r.lat_initial or r.lat_incident or r.lat_destination
    )

    def _yes(val: str | None) -> bool:
        return (val or "").strip().lower() in ("yes", "probable / inferred", "probable", "inferred")

    overview = {
        "case_count": total,
        "date_earliest": min(dates) if dates else None,
        "date_latest": max(dates) if dates else None,
        "top_cities": [{"city": c, "count": n} for c, n in cities_ctr.most_common(5)],
        "location_type_dist": [{"type": k, "count": v} for k, v in location_types.most_common()],
        "coded_count": sum(1 for r in reports if r.coding_status in ("coded", "reviewed")),
        "geocoded_count": geocoded_count,
        "harm_counts": {
            "coercion":       sum(1 for r in reports if _yes(r.coercion_present)),
            "physical_force": sum(1 for r in reports if _yes(r.physical_force)),
            "sexual_assault": sum(1 for r in reports if _yes(r.sexual_assault)),
            "robbery_theft":  sum(1 for r in reports if _yes(r.robbery_theft)),
            "threats":        sum(1 for r in reports if _yes(r.threats_present)),
            "weapon":         sum(1 for r in reports if _yes(r.weapon_present_used)),
            "restraint":      sum(1 for r in reports if _yes(r.restraint_confinement)),
            "choking":        sum(1 for r in reports if _yes(r.choking_strangulation)),
        },
    }

    map_points = [
        {
            "report_id": r.report_id,
            "lat_initial": r.lat_initial,
            "lon_initial": r.lon_initial,
            "lat_incident": r.lat_incident,
            "lon_incident": r.lon_incident,
            "lat_destination": r.lat_destination,
            "lon_destination": r.lon_destination,
            "coercion": r.coercion_present,
            "movement": r.movement_present,
            "city": r.city,
        }
        for r in reports
        if r.lat_initial or r.lat_incident
    ]

    seq_data = aggregate_sequences(reports)
    behavioral = {
        "top_sequences": seq_data["most_common_sequences"][:5],
        "escalation_points": Counter(r.escalation_point for r in reports if r.escalation_point).most_common(5),
        "top_transitions": seq_data["most_common_bigrams"][:5],
    }

    env_data = aggregate_environment(reports)

    # Stage-level situational aggregates
    from collections import defaultdict
    report_ids = [r.report_id for r in reports]
    stages_q = db.query(ReportStage).filter(ReportStage.report_id.in_(report_ids)).all() if report_ids else []
    _sit_fields = ("visibility", "guardianship", "isolation_level", "control_type")
    sit_overall: dict = {f: Counter() for f in _sit_fields}
    sit_by_stage: dict = {f: defaultdict(Counter) for f in _sit_fields}
    for s in stages_q:
        t = s.stage_type or "unknown"
        for f in _sit_fields:
            val = getattr(s, f, None)
            if val:
                sit_overall[f][val] += 1
                sit_by_stage[f][t][val] += 1
    # Collapse to plain dicts; by_stage → {stage_type: {field: top_value, field_count: n}}
    sit_by_stage_summary: dict = {}
    for stype in ["initial_contact", "negotiation", "movement", "escalation", "outcome"]:
        row: dict = {}
        for f in _sit_fields:
            top = sit_by_stage[f][stype].most_common(1)
            if top:
                row[f] = top[0][0]
                row[f + "_count"] = top[0][1]
        if row:
            sit_by_stage_summary[stype] = row

    conditions = {
        "indoor_outdoor": env_data["indoor_outdoor"],
        "public_private": env_data["public_private"],
        "deserted": env_data["deserted"],
        "location_types": env_data["location_types"][:8],
        "visibility": dict(sit_overall["visibility"]),
        "guardianship": dict(sit_overall["guardianship"]),
        "isolation_level": dict(sit_overall["isolation_level"]),
        "control_type": dict(sit_overall["control_type"]),
        "situational_by_stage": sit_by_stage_summary,
        "total_stages_coded": len(stages_q),
    }

    mob_data = aggregate_mobility(reports)
    mob_total = mob_data["total"] or 1
    movement = {
        "pct_movement": round(mob_data["counts"]["movement_present"] / mob_total * 100, 1),
        "pct_entered_vehicle": round(mob_data["counts"]["entered_vehicle"] / mob_total * 100, 1),
        "pct_public_to_private": round(mob_data["counts"]["public_to_private"] / mob_total * 100, 1),
        "top_transitions": mob_data["route_patterns"][:5],
        "common_pathways": mob_data["recurring_pathways"][:5],
    }

    # Linkage signals from the full dataset (not filtered — analysts want patterns across all cases)
    plate_ctr: Counter = Counter(r.plate_partial for r in reports if r.plate_partial)
    repeated_plates = [{"descriptor": p, "count": c} for p, c in plate_ctr.most_common(5) if c >= 2]

    make_colour_ctr: Counter = Counter()
    for r in reports:
        m = (r.vehicle_make or "").strip().title()
        cl = (r.vehicle_colour or "").strip().title()
        if m and cl:
            make_colour_ctr[f"{cl} {m}"] += 1
    repeated_make_colour = [{"descriptor": k, "count": v} for k, v in make_colour_ctr.most_common(5) if v >= 2]

    loc_ctr: Counter = Counter()
    for r in reports:
        for lf in ["initial_contact_location", "incident_location_primary"]:
            loc = (getattr(r, lf, None) or "").strip()
            if loc and len(loc) > 3:
                loc_ctr[loc] += 1
    repeated_locations = [{"descriptor": k, "count": v} for k, v in loc_ctr.most_common(5) if v >= 2]

    linkage = {
        "repeated_plates": repeated_plates,
        "repeated_vehicles": repeated_make_colour,
        "repeated_locations": repeated_locations,
        "note": "Flagged as potential linkage only — not confirmed.",
    }

    return {
        "meta": {"case_count": total, "date_from": date_from, "date_to": date_to, "status": status, "city": city},
        "overview": overview,
        "map_points": map_points,
        "behavioral": behavioral,
        "conditions": conditions,
        "movement": movement,
        "linkage": linkage,
    }


@app.get("/reports/{report_id}/summary")
def get_case_summary(report_id: str, db: Session = Depends(get_db)):
    """Case-level structured analytical summary derived from coded fields."""
    from research import build_full_case_summary
    r = db.query(Report).filter(Report.report_id == report_id).first()
    if not r:
        raise HTTPException(404, "Report not found")
    return build_full_case_summary(r)


@app.get("/export/case-summaries")
def export_case_summaries(db: Session = Depends(get_db)):
    """Export per-case analytical summaries as CSV (research-ready)."""
    from research import build_full_case_summary
    reports = db.query(Report).all()

    output = io.StringIO()
    fieldnames = [
        'report_id', 'coding_status', 'incident_date', 'city',
        'encounter_sequence', 'encounter_sequence_with_provenance',
        'has_provisional', 'mobility_summary', 'environment_summary',
        'harm_indicators', 'exit_outcome',
    ]
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()

    for r in reports:
        summary = build_full_case_summary(r)

        def _items_to_str(items):
            return '; '.join(
                item['item'] + (' [provisional]' if item['provenance'] == 'provisional' else '')
                for item in items
            )

        writer.writerow({
            'report_id':                        r.report_id,
            'coding_status':                    r.coding_status or '',
            'incident_date':                    r.incident_date or '',
            'city':                             r.city or '',
            'encounter_sequence':               summary['encounter_sequence_string'],
            'encounter_sequence_with_provenance': summary['encounter_sequence_with_provenance'],
            'has_provisional':                  'yes' if summary['has_provisional'] else 'no',
            'mobility_summary':                 _items_to_str(summary['mobility_summary']),
            'environment_summary':              _items_to_str(summary['environment_summary']),
            'harm_indicators':                  _items_to_str(summary['harm_summary']),
            'exit_outcome':                     _items_to_str(summary['exit_summary']),
        })

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type='text/csv',
        headers={'Content-Disposition': 'attachment; filename=redlight_case_summaries.csv'},
    )


@app.get("/export/research-tables")
def export_research_tables(db: Session = Depends(get_db)):
    """
    Export all research aggregate tables as a ZIP of CSVs.

    Contents:
      aggregate_sequences.csv           — full encounter sequence frequencies
      aggregate_sequence_patterns.csv   — stage-pair bigram frequencies
      stage_frequency.csv               — individual stage occurrence counts
      escalation_pathways.csv           — harm-stage-only pathway sequences
      per_case_sequences.csv            — each case's derived sequence
      aggregate_mobility_counts.csv     — mobility indicator counts + pct
      aggregate_mobility_pathways.csv   — recurring mobility combinations
      aggregate_route_patterns.csv      — start→destination type patterns
      aggregate_environment.csv         — indoor/outdoor, public/private, deserted
      aggregate_location_types.csv      — location type frequency
      aggregate_env_violence.csv        — violence/movement cross-tabulations
      aggregate_environment_patterns.csv — combined environment+harm patterns
    """
    import zipfile
    from research import aggregate_sequences, aggregate_mobility, aggregate_environment

    reports = db.query(Report).all()
    seq_data = aggregate_sequences(reports)
    mob_data = aggregate_mobility(reports)
    env_data = aggregate_environment(reports)

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:

        def _csv_str(fieldnames, rows):
            buf = io.StringIO()
            w = csv.DictWriter(buf, fieldnames=fieldnames)
            w.writeheader()
            w.writerows(rows)
            return buf.getvalue()

        def _csv_rows(headers, rows):
            buf = io.StringIO()
            w = csv.writer(buf)
            w.writerow(headers)
            w.writerows(rows)
            return buf.getvalue()

        # Sequence tables
        zf.writestr('aggregate_sequences.csv',
            _csv_str(['sequence', 'count'], seq_data['most_common_sequences']))
        zf.writestr('aggregate_sequence_patterns.csv',
            _csv_str(['pattern', 'count'], seq_data['most_common_bigrams']))
        zf.writestr('stage_frequency.csv',
            _csv_str(['stage', 'count'], seq_data['stage_frequency']))
        zf.writestr('escalation_pathways.csv',
            _csv_str(['pathway', 'count'], seq_data['escalation_pathways']))
        zf.writestr('per_case_sequences.csv',
            _csv_str(['report_id', 'sequence', 'stage_count',
                      'primary_incident_type', 'overall_severity',
                      'stage_coding_suitability', 'movement_relocation_present',
                      'escalation_cue', 'main_harms'],
                     seq_data['per_case']))

        # Mobility tables
        total_m = mob_data['total'] or 1
        mob_count_rows = [
            [k, v, round(v / total_m * 100, 1)]
            for k, v in mob_data['counts'].items()
        ]
        zf.writestr('aggregate_mobility_counts.csv',
            _csv_rows(['mobility_indicator', 'count', 'pct_of_total'], mob_count_rows))
        zf.writestr('aggregate_mobility_pathways.csv',
            _csv_str(['pathway', 'count'], mob_data['recurring_pathways']))
        zf.writestr('aggregate_route_patterns.csv',
            _csv_str(['route', 'count'], mob_data['route_patterns']))
        zf.writestr('cross_city_pathways.csv',
            _csv_str(['pathway', 'count'], mob_data['cross_city_pathways']))

        # Environment tables
        env_dist_rows = []
        for val, cnt in env_data['indoor_outdoor'].items():
            env_dist_rows.append(['indoor_outdoor', val, cnt])
        for val, cnt in env_data['public_private'].items():
            env_dist_rows.append(['public_private', val, cnt])
        for val, cnt in env_data['deserted'].items():
            env_dist_rows.append(['deserted', val, cnt])
        zf.writestr('aggregate_environment.csv',
            _csv_rows(['dimension', 'value', 'count'], env_dist_rows))

        zf.writestr('aggregate_location_types.csv',
            _csv_str(['type', 'count'], env_data['location_types']))

        # Cross-tabulations
        xtab_rows = []
        for env_dim, cross_dict in [
            ('indoor_outdoor', env_data['violence_by_environment']),
            ('public_private',  env_data['movement_by_setting']),
            ('deserted',        env_data['deserted_analysis']),
        ]:
            for val, metrics in cross_dict.items():
                xtab_rows.append([
                    env_dim, val,
                    metrics['count'],
                    metrics['physical_force'],
                    metrics['sexual_assault'],
                    metrics['coercion'],
                    metrics['movement'],
                ])
        zf.writestr('aggregate_env_violence.csv',
            _csv_rows(
                ['env_dimension', 'env_value', 'n_cases',
                 'physical_force', 'sexual_assault', 'coercion', 'movement'],
                xtab_rows,
            ))

        zf.writestr('aggregate_environment_patterns.csv',
            _csv_str(['pattern', 'count'], env_data['combined_patterns']))

        # ── Stage-level exports ───────────────────────────────────────────────
        all_stages = db.query(ReportStage).order_by(ReportStage.report_id, ReportStage.stage_order).all()
        stage_rows = []
        for s in all_stages:
            r = next((x for x in reports if x.report_id == s.report_id), None)
            stage_rows.append({
                'report_id':            s.report_id,
                'stage_order':          s.stage_order,
                'stage_type':           s.stage_type or '',
                'client_behaviors':     ', '.join(s.client_behaviors or []),
                'worker_responses':     ', '.join(s.victim_responses or []),
                'escalation_level':     getattr(s, 'escalation_level', '') or '',
                'location_type':        s.location_type or '',
                'location_label':       s.location_label or '',
                'movement_type_to_here':s.movement_type_to_here or '',
                'movement_impact':      getattr(s, 'movement_impact', '') or '',
                'visibility':           s.visibility or '',
                'guardianship':         s.guardianship or '',
                'isolation_level':      s.isolation_level or '',
                'control_type':         s.control_type or '',
                'able_to_leave':        getattr(s, 'able_to_leave', '') or '',
                'spatial_precision':    getattr(s, 'spatial_precision', '') or '',
                'turning_point_notes':  s.turning_point_notes or '',
                'supporting_excerpt':   getattr(s, 'supporting_excerpt', '') or '',
                'coder_notes_stage':    getattr(s, 'coder_notes_stage', '') or '',
                'coding_confidence':    getattr(s, 'coding_confidence', '') or '',
                'analyst_name':         (r.analyst_name if r else '') or '',
                'coding_status':        (r.coding_status if r else '') or '',
                'incident_date':        (r.incident_date if r else '') or '',
                'provenance':           'analyst_confirmed',
            })

        _STAGE_FIELDS = ['report_id','stage_order','stage_type','client_behaviors','worker_responses',
            'escalation_level','location_type','location_label','movement_type_to_here','movement_impact',
            'visibility','guardianship','isolation_level','control_type','able_to_leave','spatial_precision',
            'turning_point_notes','supporting_excerpt','coder_notes_stage','coding_confidence',
            'analyst_name','coding_status','incident_date','provenance']
        zf.writestr('stage_sequences.csv', _csv_str(_STAGE_FIELDS, stage_rows))

        # Stage transitions (consecutive stage pairs per report)
        transition_rows = []
        from itertools import groupby as _groupby
        from operator import attrgetter as _attrgetter
        for report_id, grp in _groupby(all_stages, key=_attrgetter('report_id')):
            grp_list = sorted(grp, key=lambda s: s.stage_order)
            r = next((x for x in reports if x.report_id == report_id), None)
            for i in range(len(grp_list) - 1):
                a, b = grp_list[i], grp_list[i + 1]
                has_mov = (a.movement_type_to_here or b.movement_type_to_here or '') not in ('', 'no_movement', 'none')
                transition_rows.append({
                    'report_id':             report_id,
                    'from_stage_order':      a.stage_order,
                    'to_stage_order':        b.stage_order,
                    'from_stage_type':       a.stage_type or '',
                    'to_stage_type':         b.stage_type or '',
                    'from_escalation_level': getattr(a, 'escalation_level', '') or '',
                    'to_escalation_level':   getattr(b, 'escalation_level', '') or '',
                    'movement_involved':     'yes' if has_mov else 'no',
                    'public_to_private_shift': 'yes' if (a.visibility in ('public','semi_public') and b.visibility in ('semi_private','private')) else 'no',
                    'isolation_increased':   'yes' if (a.isolation_level in ('not_isolated','partially_isolated') and b.isolation_level == 'isolated') else 'no',
                    'movement_impact':       getattr(b, 'movement_impact', '') or '',
                    'incident_date':         (r.incident_date if r else '') or '',
                })
        _TRANS_FIELDS = ['report_id','from_stage_order','to_stage_order','from_stage_type','to_stage_type',
            'from_escalation_level','to_escalation_level','movement_involved','public_to_private_shift',
            'isolation_increased','movement_impact','incident_date']
        zf.writestr('stage_transitions.csv', _csv_str(_TRANS_FIELDS, transition_rows))

    zip_buf.seek(0)
    return StreamingResponse(
        iter([zip_buf.getvalue()]),
        media_type='application/zip',
        headers={'Content-Disposition': 'attachment; filename=redlight_research_tables.zip'},
    )


@app.get("/export/methodology-summary")
def export_methodology_summary(db: Session = Depends(get_db)):
    """
    Export a coding coverage report for dissertation methodology documentation.
    Produces a CSV suitable for pasting into a methods chapter or appendix.
    """
    from datetime import datetime, timezone
    reports = db.query(Report).all()
    total = len(reports)

    _POS = {'yes', 'probable', 'inferred', 'probable / inferred'}

    def _has_analyst_field(r, *fields):
        fp = getattr(r, 'field_provenance', None) or {}
        for f in fields:
            val = (getattr(r, f, None) or '').strip()
            prov = fp.get(f, 'unset') if isinstance(fp, dict) else 'unset'
            if val and prov in ('analyst_filled', 'reviewed'):
                return True
        return False

    _HARM_FIELDS = ['coercion_present','threats_present','physical_force','sexual_assault',
                    'stealthing','robbery_theft','forced_movement_dragging','restraint_confinement',
                    'weapon_present_used','choking_strangulation','prevented_exit',
                    'non_consensual_substance','loss_of_consciousness']
    _MOV_FIELDS  = ['movement_present','mode_of_movement','entered_vehicle',
                    'public_to_private_shift','cross_municipality']
    _SAF_FIELDS  = ['concern_trafficking','concern_third_party_control','concern_grooming',
                    'concern_organized_offending','concern_urgent_public_safety']

    stage_report_ids = {s.report_id for s in db.query(ReportStage.report_id).distinct().all()}

    nlp_fields = ['nlp_coercion_rank','nlp_physical_rank','nlp_sexual_rank','nlp_movement_rank']

    rows = [
        ('Metric', 'Count', 'Pct of imported', 'Notes'),
        ('Reports imported', total, '100%', 'Total source reports in database'),
        ('Reports NLP screened', total, '100%', 'All reports pass NLP extraction on import — signals are provisional'),
        ('Reports analyst coded', sum(1 for r in reports if r.coding_status in ('coded','reviewed')), f'{round(sum(1 for r in reports if r.coding_status in ("coded","reviewed"))/total*100,1) if total else 0}%', 'coding_status = coded or reviewed'),
        ('Reports analyst staged', len(stage_report_ids), f'{round(len(stage_report_ids)/total*100,1) if total else 0}%', 'Reports with analyst-created stage records in stages table'),
        ('Reports geocoded', sum(1 for r in reports if (r.lat_initial or r.lat_incident)), f'{round(sum(1 for r in reports if (r.lat_initial or r.lat_incident))/total*100,1) if total else 0}%', 'At least one geocoded point (initial or incident)'),
        ('Reports with location field coded', sum(1 for r in reports if (r.initial_contact_location or r.incident_location_primary or '').strip()), '', 'Has text in initial_contact_location or incident_location_primary'),
        ('Reports with harm fields coded', sum(1 for r in reports if any((getattr(r,f,'') or '') in _POS for f in _HARM_FIELDS)), '', 'At least one harm indicator in positive set'),
        ('Reports with movement coded', sum(1 for r in reports if (r.movement_present or '').strip()), '', 'movement_present field has any value'),
        ('Reports with public safety flags coded', sum(1 for r in reports if any((getattr(r,f,'') or '') in _POS for f in _SAF_FIELDS)), '', 'At least one concern flag in positive set'),
        ('Reports with NLP provisional signals', sum(1 for r in reports if any(getattr(r, f, 0) or 0 for f in nlp_fields)), '', 'NLP rank score > 0 for any signal — provisional, pending analyst confirmation'),
        ('Reports needing geocode (location phrase present, no coordinates)', sum(1 for r in reports if (r.initial_contact_location or r.incident_location_primary or '').strip() and not (r.lat_initial or r.lat_incident)), '', ''),
        ('', '', '', ''),
        ('Export timestamp', datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC'), '', ''),
        ('NLP note', 'NLP signals are provisional — not treated as findings until analyst confirmed', '', ''),
        ('Auditability note', 'All coded observations are analyst observations only, not confirmed legal or investigative findings', '', ''),
    ]

    output = io.StringIO()
    w = csv.writer(output)
    for row in rows:
        w.writerow(row)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type='text/csv',
        headers={'Content-Disposition': 'attachment; filename=coding_coverage_report.csv'},
    )


@app.get("/export/codebook")
def export_codebook(db: Session = Depends(get_db)):
    """
    Export a codebook listing all coding fields, allowed values, and provenance notes.
    Useful for dissertation methodology appendix.
    """
    CODEBOOK: list[tuple] = [
        # (field_name, tab/section, definition, allowed_values, source, usable_in_findings)

        # ── Admin / Identity ────────────────────────────────────────────────
        ('report_id',              'Admin', 'Unique case identifier (UUID)', 'UUID string', 'system_generated', 'no'),
        ('coding_status',          'Admin', 'Current coding stage of this report', 'uncoded | in_progress | coded | reviewed', 'analyst', 'yes'),
        ('analyst_name',           'Admin', 'Name of analyst who coded the report', 'free text', 'analyst', 'yes'),
        ('date_received',          'Admin', 'Date the source report was imported', 'YYYY-MM-DD', 'import', 'no'),
        ('source_organization',    'Admin', 'Organisation that produced the source report', 'free text', 'import', 'yes'),
        ('confidence_level',       'Admin', 'Overall analyst confidence in coding accuracy', 'low | medium | high', 'analyst', 'yes'),

        # ── Basics ──────────────────────────────────────────────────────────
        ('incident_date',          'Basics', 'Date the incident occurred', 'YYYY-MM-DD or free text', 'analyst', 'yes'),
        ('incident_time_range',    'Basics', 'Approximate time period of the incident', 'free text', 'analyst', 'yes'),
        ('day_of_week',            'Basics', 'Day of week of the incident', 'Monday–Sunday | unknown', 'analyst', 'yes'),
        ('city',                   'Basics', 'Primary city label for the case (legacy summary field)', 'free text', 'analyst', 'yes'),
        ('neighbourhood',          'Basics', 'Neighbourhood or area within the city', 'free text', 'analyst', 'yes'),
        ('initial_contact_city',          'Basics', 'City where initial contact occurred', 'free text', 'analyst', 'yes'),
        ('initial_contact_city_confidence', 'Basics', 'Confidence level for initial contact city', 'known | probable | inferred | unknown', 'analyst', 'yes'),
        ('incident_city',          'Basics', 'City where primary incident occurred', 'free text', 'analyst', 'yes'),
        ('incident_city_confidence', 'Basics', 'Confidence level for incident city', 'known | probable | inferred | unknown', 'analyst', 'yes'),
        ('destination_city',       'Basics', 'City at destination/secondary location', 'free text', 'analyst', 'yes'),
        ('destination_city_confidence', 'Basics', 'Confidence level for destination city', 'known | probable | inferred | unknown', 'analyst', 'yes'),
        ('cross_city_movement',    'Basics', 'Whether movement crossed city boundaries', 'yes | no | unclear', 'analyst', 'yes'),
        ('initial_contact_location',  'Basics', 'Textual description of initial contact location', 'free text', 'analyst', 'yes'),
        ('incident_location_primary', 'Basics', 'Textual description of primary incident location', 'free text', 'analyst', 'yes'),
        ('indoor_outdoor',         'Basics', 'Whether incident took place indoors or outdoors', 'indoor | outdoor | both | unknown', 'analyst', 'yes'),
        ('public_private',         'Basics', 'Public or private setting', 'public | semi-public | private | mixed | unknown', 'analyst', 'yes'),
        ('deserted',               'Basics', 'Whether the location was deserted / isolated', 'deserted | not_deserted | unclear', 'analyst', 'yes'),

        # ── Encounter / Incident Overview ───────────────────────────────────
        ('primary_incident_type',    'Encounter / Overview', 'Most analytically significant harm type for the whole incident', 'suspicious / concerning behaviour | non-payment / payment dispute | coercion / intimidation | physical violence | sexual violence | robbery / theft | substance-facilitated harm | movement / relocation concern | multiple harms | other | unknown / unclear', 'analyst', 'yes'),
        ('overall_severity',         'Encounter / Overview', 'Analyst assessment of overall incident severity', 'low concern | moderate concern | high concern | severe violence / high risk | unknown / unclear', 'analyst', 'yes'),
        ('overall_incident_summary', 'Encounter / Overview', 'Brief analyst narrative summary of the whole incident', 'free text', 'analyst', 'yes'),
        ('stage_coding_suitability', 'Encounter / Overview', 'Whether narrative supports full stage-by-stage coding', 'Yes, sufficient narrative detail for staged coding | Partial, some stages can be coded | No, incident-level coding only | Unknown / not reviewed', 'analyst', 'yes'),
        ('sequence_clarity',         'Encounter / Overview', 'Clarity of the described event sequence', 'clear linear sequence | mostly clear, minor gaps | fragmented / partially inferrable | too fragmented to sequence | unknown', 'analyst', 'yes'),
        ('boundary_issue_present',   'Encounter / Overview', 'Whether a consent-boundary violation is clearly present', 'yes | no | unclear | unknown', 'analyst', 'yes'),
        ('movement_relocation_present', 'Encounter / Overview', 'Whether movement or relocation is a key feature of this case', 'yes | no | unclear | unknown', 'analyst', 'yes'),
        ('key_supporting_excerpts',  'Encounter / Overview', 'Verbatim text from source that anchors the overall coding', 'free text (verbatim quotes)', 'analyst', 'yes'),

        # ── Encounter / Harm Indicators ─────────────────────────────────────
        ('initial_approach_type',    'Encounter / Harm', 'How initial contact was made', 'street | online | referral | phone | other | unclear', 'analyst', 'yes'),
        ('negotiation_present',      'Encounter / Harm', 'Whether a negotiation phase occurred', 'yes | no | unclear', 'analyst', 'yes'),
        ('refusal_present',          'Encounter / Harm', 'Whether the worker refused a request at any point', 'yes | no | unclear', 'analyst', 'yes'),
        ('pressure_after_refusal',   'Encounter / Harm', 'Whether pressure was applied after refusal', 'yes | no | unclear', 'analyst', 'yes'),
        ('coercion_present',         'Encounter / Harm', 'Whether coercive behaviour was present anywhere in the incident', 'yes | no | unclear / not stated | probable / inferred | not coded', 'analyst', 'yes'),
        ('threats_present',          'Encounter / Harm', 'Whether explicit or implicit threats were made', 'yes | no | unclear / not stated | probable / inferred | not coded', 'analyst', 'yes'),
        ('verbal_abuse',             'Encounter / Harm', 'Whether verbal abuse occurred', 'yes | no | unclear / not stated | probable / inferred | not coded', 'analyst', 'yes'),
        ('physical_force',           'Encounter / Harm', 'Whether physical force was used', 'yes | no | unclear / not stated | probable / inferred | not coded', 'analyst', 'yes'),
        ('sexual_assault',           'Encounter / Harm', 'Whether sexual assault occurred', 'yes | no | unclear / not stated | probable / inferred | not coded', 'analyst', 'yes'),
        ('robbery_theft',            'Encounter / Harm', 'Whether robbery or theft occurred', 'yes | no | unclear / not stated | probable / inferred | not coded', 'analyst', 'yes'),
        ('stealthing',               'Encounter / Harm', 'Whether stealthing (non-consensual removal of condom) occurred', 'yes | no | unclear / not stated | probable / inferred | not coded', 'analyst', 'yes'),
        ('forced_movement_dragging', 'Encounter / Harm', 'Whether forced physical movement / dragging occurred', 'yes | no | unclear / not stated | probable / inferred | not coded', 'analyst', 'yes'),
        ('restraint_confinement',    'Encounter / Harm', 'Whether restraint or physical confinement occurred', 'yes | no | unclear / not stated | probable / inferred | not coded', 'analyst', 'yes'),
        ('weapon_present_used',      'Encounter / Harm', 'Whether a weapon was present or used', 'yes | no | unclear / not stated | probable / inferred | not coded', 'analyst', 'yes'),
        ('choking_strangulation',    'Encounter / Harm', 'Whether choking or strangulation occurred', 'yes | no | unclear / not stated | probable / inferred | not coded', 'analyst', 'yes'),
        ('prevented_exit',           'Encounter / Harm', 'Whether the worker was prevented from leaving', 'yes | no | unclear / not stated | probable / inferred | not coded', 'analyst', 'yes'),
        ('non_consensual_substance', 'Encounter / Harm', 'Whether a substance was administered without consent', 'yes | no | unclear / not stated | probable / inferred | not coded', 'analyst', 'yes'),
        ('loss_of_consciousness',    'Encounter / Harm', 'Whether loss of consciousness occurred', 'yes | no | unclear / not stated | probable / inferred | not coded', 'analyst', 'yes'),
        ('exit_type',                'Encounter / Harm', 'How the incident ended or was resolved', 'completed | escaped | abandoned | interrupted | unknown', 'analyst', 'yes'),
        ('repeated_pressure',        'Encounter / Harm', 'Whether repeated pressure / persistence was documented', 'yes | no | unclear', 'analyst', 'yes'),
        ('intimidation_present',     'Encounter / Harm', 'Whether intimidation was present (short of explicit threats)', 'yes | no | unclear', 'analyst', 'yes'),
        ('abrupt_tone_change',       'Encounter / Harm', 'Whether an abrupt change in client tone or behaviour was noted', 'yes | no | unclear', 'analyst', 'yes'),

        # ── VAWG / Exploitation flags ────────────────────────────────────────
        ('trafficking_exploitation_concern',    'VAWG / Exploitation', 'Analyst concern: trafficking or exploitation indicators present', 'yes | no | unclear | probable | inferred | unknown', 'analyst', 'yes — observation only'),
        ('third_party_control_indicated',       'VAWG / Exploitation', 'Analyst concern: evidence suggesting a third party controls the worker', 'yes | no | unclear | probable | inferred | unknown', 'analyst', 'yes — observation only'),
        ('worker_appears_controlled',           'VAWG / Exploitation', 'Worker behavioural indicators suggesting control or coercion by another party', 'yes | no | unclear | probable | inferred | unknown', 'analyst', 'yes — observation only'),
        ('client_connected_to_controller',      'VAWG / Exploitation', 'Indication that the client may be connected to a controller or manager', 'yes | no | unclear | probable | inferred | unknown', 'analyst', 'yes — observation only'),
        ('movement_to_unknown_unsafe_location', 'VAWG / Exploitation', 'Worker was moved to a location she did not know or that appeared unsafe', 'yes | no | unclear | probable | inferred | unknown', 'analyst', 'yes — observation only'),
        ('worker_unaware_how_arrived',          'VAWG / Exploitation', 'Worker did not appear to know how she arrived at the location', 'yes | no | unclear | probable | inferred | unknown', 'analyst', 'yes — observation only'),
        ('grooming_recruitment_concern',        'VAWG / Exploitation', 'Indicators suggestive of grooming or recruitment dynamics', 'yes | no | unclear | probable | inferred | unknown', 'analyst', 'yes — observation only'),
        ('repeat_targeting_concern',            'VAWG / Exploitation', 'Evidence or pattern suggesting this worker is repeat-targeted', 'yes | no | unclear | probable | inferred | unknown', 'analyst', 'yes — observation only'),
        ('multiple_women_referenced',           'VAWG / Exploitation', 'Report references multiple women, suggesting possible organised exploitation', 'yes | no | unclear | probable | inferred | unknown', 'analyst', 'yes — observation only'),
        ('organized_group_offending_concern',   'VAWG / Exploitation', 'Indicators of organised or group-based offending', 'yes | no | unclear | probable | inferred | unknown', 'analyst', 'yes — observation only'),
        ('public_safety_bulletin_suitability',  'VAWG / Exploitation', 'Whether this case warrants a public safety bulletin', 'yes | no | unclear | unknown', 'analyst', 'yes'),
        ('public_safety_urgency_level',         'VAWG / Exploitation', 'Urgency level for public safety action', 'immediate | high | moderate | low | unknown', 'analyst', 'yes'),
        ('vawg_exploitation_notes',             'VAWG / Exploitation', 'Analyst notes on VAWG/exploitation indicators', 'free text', 'analyst', 'yes'),
        ('vawg_key_excerpts',                   'VAWG / Exploitation', 'Verbatim source text excerpts supporting VAWG/exploitation coding', 'free text (verbatim quotes)', 'analyst', 'yes'),

        # ── Mobility ────────────────────────────────────────────────────────
        ('movement_present',              'Mobility', 'Whether physical movement or relocation occurred', 'yes | no | unclear | unknown', 'analyst', 'yes'),
        ('movement_attempted',            'Mobility', 'Whether movement was attempted but did not occur', 'yes | no | unclear', 'analyst', 'yes'),
        ('mode_of_movement',              'Mobility', 'Primary mode of transport if movement occurred', 'on foot | vehicle | taxi/rideshare | other | unknown', 'analyst', 'yes'),
        ('entered_vehicle',               'Mobility', 'Whether the worker entered a vehicle', 'yes | no | unclear | unknown', 'analyst', 'yes'),
        ('vehicle_driver_role',           'Mobility', 'Role of vehicle driver in the incident', 'free text', 'analyst', 'yes'),
        ('start_location_type',           'Mobility', 'Environment type at movement start', 'street_outdoor | vehicle | hotel_motel | residence | bar_club | other | unknown', 'analyst', 'yes'),
        ('destination_location_type',     'Mobility', 'Environment type at movement destination', 'street_outdoor | vehicle | hotel_motel | residence | bar_club | other | unknown', 'analyst', 'yes'),
        ('public_to_private_shift',       'Mobility', 'Whether movement involved a shift from public to private space', 'yes | no | unclear | unknown', 'analyst', 'yes'),
        ('public_to_secluded_shift',      'Mobility', 'Whether movement involved a shift to a secluded location', 'yes | no | unclear | unknown', 'analyst', 'yes'),
        ('cross_neighbourhood',           'Mobility', 'Whether movement crossed neighbourhood boundaries', 'yes | no | unclear | unknown', 'analyst', 'yes'),
        ('cross_municipality',            'Mobility', 'Whether movement crossed municipal boundaries', 'yes | no | unclear | unknown', 'analyst', 'yes'),
        ('offender_control_over_movement','Mobility', 'Degree of offender control over movement direction/destination', 'low | moderate | high | unclear', 'analyst', 'yes'),
        ('movement_completed',            'Mobility', 'Whether the described movement was completed', 'yes | no | unclear', 'analyst', 'yes'),
        ('who_controlled_movement',       'Mobility', 'Who had primary control over movement', 'offender | victim | shared | unclear', 'analyst', 'yes'),
        ('unexplained_relocation',        'Mobility', 'Whether an unexplained or sudden relocation occurred', 'yes | no | unclear', 'analyst', 'yes'),
        ('movement_confidence',           'Mobility', 'Analyst confidence in movement coding', 'high | medium | low | unclear', 'analyst', 'yes'),
        ('movement_purpose',              'Mobility', 'Inferred or stated purpose of the movement', 'relocation to more private setting | movement during escape | offender-controlled transport | victim-controlled movement | unknown | other', 'analyst', 'yes'),
        ('basis_for_movement_coding',     'Mobility', 'What the movement coding is based on', 'explicit narrative | inferred from sequence | inferred from GIS | analyst judgement | other', 'analyst', 'yes'),
        ('destination_known',             'Mobility', 'Whether the destination was known to the worker before movement', 'yes | no | unclear | inferred', 'analyst', 'yes'),
        ('location_certainty',            'Mobility', 'Overall certainty in location coding', 'high | medium | low | unknown', 'analyst', 'yes'),

        # ── Harm Classification ──────────────────────────────────────────────
        ('primary_harm',    'Harm Classification', 'Single most analytically significant harm category coded for this case', 'coercion / intimidation | physical violence | sexual violence | robbery / theft | substance-facilitated harm | movement / relocation concern | suspicious / concerning behaviour | multiple harms | other | unknown', 'analyst', 'yes'),
        ('multi_harm_flag', 'Harm Classification', 'Whether more than one harm category was coded for this case', 'yes | no | unclear', 'analyst', 'yes'),

        # ── Narrative / Summary Coding ───────────────────────────────────────
        ('early_escalation_score',     'Narrative', 'Composite score reflecting early escalation indicators (0–10)', '0–10 numeric', 'analyst', 'yes'),
        ('escalation_point',           'Narrative', 'Point in the encounter where escalation was first identified', 'free text', 'analyst', 'yes'),
        ('highest_stage_reached',      'Narrative', 'Highest level of harm or escalation recorded for this case', 'no clear escalation | negotiation conflict | coercion/control | physical violence | sexual violence | robbery/theft | mixed severe harm | unknown', 'analyst', 'yes'),
        ('turning_point',              'Narrative', 'Key transition moment in the encounter', 'boundary tested | refusal ignored | pressure increased | deception | movement imposed | isolation increased | threat introduced | exit blocked | physical force applied | sexual violence initiated | robbery initiated | other', 'analyst', 'yes'),
        ('resolution_endpoint',        'Narrative', 'How the incident concluded', 'victim escaped | offender left | assault completed | robbery completed | third-party interruption | unknown | other', 'analyst', 'yes'),
        ('summary_analytic',           'Narrative', 'Analyst interpretive summary', 'free text', 'analyst', 'yes'),
        ('key_quotes',                 'Narrative', 'Key verbatim quotes from source supporting coding', 'free text (verbatim quotes)', 'analyst', 'yes'),
        ('coder_notes',                'Narrative', 'General coder notes and uncertainties', 'free text', 'analyst', 'no — internal working notes'),
        ('analyst_summary',            'Narrative', 'Analyst interpretive narrative (distinct from cleaned transcript)', 'free text', 'analyst', 'yes'),
        ('cleaned_narrative',          'Narrative', 'Cleaned / redacted version of the source narrative', 'free text', 'analyst', 'no — internal only'),

        # ── Suspect / Vehicle ────────────────────────────────────────────────
        ('suspect_count',                   'Suspect/Vehicle', 'Number of suspects involved', 'numeric or free text', 'analyst', 'yes'),
        ('suspect_gender',                  'Suspect/Vehicle', 'Gender of primary suspect as reported', 'free text', 'analyst/nlp', 'yes'),
        ('suspect_race_ethnicity',          'Suspect/Vehicle', 'Race/ethnicity of suspect as reported in source', 'free text', 'analyst', 'no — descriptor only; treat with methodological caution'),
        ('suspect_age_estimate',            'Suspect/Vehicle', 'Estimated age or age range of suspect', 'free text', 'analyst', 'yes'),
        ('suspect_distinctive_features',    'Suspect/Vehicle', 'Distinctive physical features of suspect', 'free text', 'analyst', 'no — descriptor only'),
        ('suspect_clothing',                'Suspect/Vehicle', 'Clothing description of suspect', 'free text', 'analyst', 'no — descriptor only'),
        ('suspect_speech_notes',            'Suspect/Vehicle', 'Notable speech, language, or accent features of suspect', 'free text', 'analyst', 'no — descriptor only'),
        ('suspect_behavioural_descriptors', 'Suspect/Vehicle', 'Behavioural descriptors for suspect', 'free text', 'analyst', 'yes'),
        ('known_repeat_suspect',            'Suspect/Vehicle', 'Whether this suspect is a known repeat offender in the dataset', 'yes | no | unclear | unknown', 'analyst', 'yes'),
        ('repeat_suspect_flag',             'Suspect/Vehicle', 'Flag: suspect has been seen in multiple reports', 'yes | no | unknown', 'analyst', 'yes'),
        ('vehicle_present',                 'Suspect/Vehicle', 'Whether a vehicle was involved', 'yes | no | unknown', 'analyst', 'yes'),
        ('vehicle_make',                    'Suspect/Vehicle', 'Make of suspect vehicle', 'free text', 'analyst', 'no — descriptor only'),
        ('vehicle_year',                    'Suspect/Vehicle', 'Year of suspect vehicle', 'free text', 'analyst', 'no — descriptor only'),
        ('vehicle_model',                   'Suspect/Vehicle', 'Model of suspect vehicle', 'free text', 'analyst', 'no — descriptor only'),
        ('vehicle_colour',                  'Suspect/Vehicle', 'Colour of suspect vehicle', 'free text', 'analyst', 'no — descriptor only'),
        ('plate_partial',                   'Suspect/Vehicle', 'Partial or full licence plate', 'free text', 'analyst', 'no — for investigative use only'),
        ('vehicle_role_in_encounter',       'Suspect/Vehicle', 'Role the vehicle played in the incident', 'transportation only | offence location | movement/control mechanism | escape vehicle | unknown | other', 'analyst', 'yes'),
        ('vehicle_ownership_association',   'Suspect/Vehicle', 'Vehicle ownership or association', 'suspect vehicle | victim/survivor vehicle | rideshare/taxi | rental | unknown', 'analyst', 'yes'),
        ('vehicle_confidence',              'Suspect/Vehicle', 'Confidence in vehicle information', 'high | medium | low | unclear', 'analyst', 'yes'),
        ('repeat_vehicle_flag',             'Suspect/Vehicle', 'Flag: vehicle has appeared in multiple reports', 'yes | no | unknown', 'analyst', 'yes'),

        # ── Concern Flags ────────────────────────────────────────────────────
        ('concern_trafficking',          'Concern Flags', 'Analyst concern: trafficking or exploitation indicators', 'yes | no | unclear | probable | inferred | unknown', 'analyst', 'yes — observation only, not confirmed finding'),
        ('concern_third_party_control',  'Concern Flags', 'Analyst concern: third-party control indicators', 'yes | no | unclear | probable | inferred | unknown', 'analyst', 'yes — observation only, not confirmed finding'),
        ('concern_grooming',             'Concern Flags', 'Analyst concern: grooming or recruitment dynamics', 'yes | no | unclear | probable | inferred | unknown', 'analyst', 'yes — observation only, not confirmed finding'),
        ('concern_organized_offending',  'Concern Flags', 'Analyst concern: organised group offending', 'yes | no | unclear | probable | inferred | unknown', 'analyst', 'yes — observation only, not confirmed finding'),
        ('concern_repeat_suspect',       'Concern Flags', 'Analyst concern: repeat suspect pattern', 'yes | no | unclear | probable | inferred | unknown', 'analyst', 'yes — observation only, not confirmed finding'),
        ('concern_repeat_vehicle',       'Concern Flags', 'Analyst concern: repeat vehicle pattern', 'yes | no | unclear | probable | inferred | unknown', 'analyst', 'yes — observation only, not confirmed finding'),
        ('concern_urgent_public_safety', 'Concern Flags', 'Analyst concern: urgent public safety risk', 'yes | no | unclear | probable | inferred | unknown', 'analyst', 'yes — observation only, not confirmed finding'),
        ('concern_bulletin_suitable',    'Concern Flags', 'Whether this case warrants a public safety bulletin issuance', 'yes | no | unclear | probable | inferred | unknown', 'analyst', 'yes — observation only, not confirmed finding'),
        ('concern_flag_rationale',       'Concern Flags', 'Analyst rationale for concern flag coding decisions', 'free text', 'analyst', 'yes'),

        # ── GIS ──────────────────────────────────────────────────────────────
        ('lat_initial',                       'GIS', 'Latitude of initial contact point (geocoded)', 'decimal degrees', 'geocoded', 'yes'),
        ('lon_initial',                       'GIS', 'Longitude of initial contact point (geocoded)', 'decimal degrees', 'geocoded', 'yes'),
        ('lat_incident',                      'GIS', 'Latitude of primary incident point (geocoded)', 'decimal degrees', 'geocoded', 'yes'),
        ('lon_incident',                      'GIS', 'Longitude of primary incident point (geocoded)', 'decimal degrees', 'geocoded', 'yes'),
        ('lat_destination',                   'GIS', 'Latitude of destination/secondary point (geocoded)', 'decimal degrees', 'geocoded', 'yes'),
        ('lon_destination',                   'GIS', 'Longitude of destination/secondary point (geocoded)', 'decimal degrees', 'geocoded', 'yes'),
        ('initial_contact_address_normalized','GIS', 'Normalised/geocoded address for initial contact point', 'free text', 'geocoded', 'yes'),
        ('initial_contact_precision',         'GIS', 'Geocoding precision for initial contact point', 'exact_address | intersection | landmark | neighbourhood | approximate | unknown', 'analyst/geocoded', 'yes'),
        ('initial_contact_source',            'GIS', 'Source basis for initial contact location', 'stated | inferred | unclear', 'analyst', 'yes'),
        ('initial_contact_confidence',        'GIS', 'Confidence in initial contact geocoding', 'high | medium | low | none', 'analyst', 'yes'),
        ('initial_contact_location_type',     'GIS', 'Environment type at initial contact point', 'street_outdoor | vehicle | hotel_motel | residence | bar_club | other | unknown', 'analyst', 'yes'),
        ('initial_contact_geocoding_status',  'GIS', 'Geocoding completion status for initial contact', 'geocoded | needs_review | not_attempted | no_location_info', 'analyst/system', 'yes'),
        ('incident_address_normalized',       'GIS', 'Normalised/geocoded address for incident point', 'free text', 'geocoded', 'yes'),
        ('incident_precision',                'GIS', 'Geocoding precision for incident point', 'exact_address | intersection | landmark | neighbourhood | approximate | unknown', 'analyst/geocoded', 'yes'),
        ('incident_source',                   'GIS', 'Source basis for incident location', 'stated | inferred | unclear', 'analyst', 'yes'),
        ('incident_confidence',               'GIS', 'Confidence in incident geocoding', 'high | medium | low | none', 'analyst', 'yes'),
        ('incident_location_type',            'GIS', 'Environment type at incident point', 'street_outdoor | vehicle | hotel_motel | residence | bar_club | other | unknown', 'analyst', 'yes'),
        ('incident_geocoding_status',         'GIS', 'Geocoding completion status for incident point', 'geocoded | needs_review | not_attempted | no_location_info', 'analyst/system', 'yes'),
        ('destination_address_normalized',    'GIS', 'Normalised/geocoded address for destination point', 'free text', 'geocoded', 'yes'),
        ('destination_precision',             'GIS', 'Geocoding precision for destination point', 'exact_address | intersection | landmark | neighbourhood | approximate | unknown', 'analyst/geocoded', 'yes'),
        ('destination_source',                'GIS', 'Source basis for destination location', 'stated | inferred | unclear', 'analyst', 'yes'),
        ('destination_confidence',            'GIS', 'Confidence in destination geocoding', 'high | medium | low | none', 'analyst', 'yes'),
        ('destination_geocoding_status',      'GIS', 'Geocoding completion status for destination point', 'geocoded | needs_review | not_attempted | no_location_info', 'analyst/system', 'yes'),

        # ── NLP (provisional — do not use as confirmed findings) ────────────
        ('nlp_coercion_rank',  'NLP (provisional)', 'NLP signal rank for coercion (0=none, 1=weak, 2=moderate, 3=strong)', '0–3 integer', 'nlp_provisional', 'NO — provisional, requires analyst confirmation'),
        ('nlp_physical_rank',  'NLP (provisional)', 'NLP signal rank for physical violence', '0–3 integer', 'nlp_provisional', 'NO — provisional, requires analyst confirmation'),
        ('nlp_sexual_rank',    'NLP (provisional)', 'NLP signal rank for sexual violence', '0–3 integer', 'nlp_provisional', 'NO — provisional, requires analyst confirmation'),
        ('nlp_movement_rank',  'NLP (provisional)', 'NLP signal rank for movement/relocation', '0–3 integer', 'nlp_provisional', 'NO — provisional, requires analyst confirmation'),
    ]

    output = io.StringIO()
    w = csv.writer(output)
    w.writerow(['field_name', 'tab_section', 'definition', 'allowed_values', 'source', 'usable_in_confirmed_findings'])
    w.writerows(CODEBOOK)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type='text/csv',
        headers={'Content-Disposition': 'attachment; filename=codebook.csv'},
    )


# ── Static files (production build) ──────────────────────────────────────────
# Serve the Vite-built frontend when it exists.
# API routes above take priority; this catches everything else.

_DIST = os.path.join(os.path.dirname(__file__), '..', 'frontend', 'dist')

if os.path.isdir(_DIST):
    app.mount("/assets", StaticFiles(directory=os.path.join(_DIST, "assets")), name="assets")

    @app.get("/{full_path:path}", include_in_schema=False)
    def spa_fallback(full_path: str):
        # Serve real files that exist in the dist root (logo.png, favicon, etc.)
        candidate = os.path.join(_DIST, full_path)
        if os.path.isfile(candidate):
            return FileResponse(candidate)
        index = os.path.join(_DIST, "index.html")
        return FileResponse(index, headers={"Cache-Control": "no-store"})
